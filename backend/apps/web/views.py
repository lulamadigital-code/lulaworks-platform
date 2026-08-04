"""Manager web dashboards (server-rendered HTML + HTMX).

Session-authenticated pages for office/manager users — the data-heavy surface that
suits a DOM web app. Deliberately separate from the JWT API (which the Flutter
field app uses). Reuses the exact same services, so there is one source of truth
for readiness, health and profitability.

Tenancy: the ambient TenantMiddleware binds the tenant from request.user for the
whole request, so `Project.objects.all()` is already tenant-scoped here.
Golden Rule: money is computed/shown only for users with `finance.view_money`.
"""

import re
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views.decorators.http import require_POST

from apps.ai_platform.decomposition import (
    apply_decomposition,
    propose_decomposition,
    record_proposal,
)
from apps.ai_platform.orchestrator import orchestrate
from apps.compliance.models import ComplianceItem
from apps.compliance.services import approve_item, recompute_readiness
from apps.compliance.services import override as override_gate
from apps.estimating.models import Estimate, EstimateStatus
from apps.estimating.services import approve_estimate, create_revision
from apps.execution.models import (
    Assignment,
    ChecklistItem,
    Phase,
    RiskLevel,
    Subtask,
    Task,
    TaskDependency,
    TaskPriority,
    TaskStatus,
    WorkOrigin,
)
from apps.execution.services import (
    add_attachment,
    add_checklist_item,
    add_comment,
    add_member,
    add_phase,
    add_subtask,
    complete_task,
    compute_task_readiness,
    create_work,
    ensure_default_phases,
    has_work_perm,
    link_tasks,
    mark_notifications_read,
    portfolio_report,
    project_health,
    remove_member,
    start_task,
    toggle_checklist_item,
    transition,
    unread_count,
    work_dashboard,
)
from apps.identity.models import (
    CompanyBankAccount,
    CompanyContact,
    CompanyDocument,
    Membership,
    Role,
)
from apps.identity.profile import (
    add_bank_account,
    add_contact,
    completeness,
    default_bank_account,
    get_profile,
    set_default_bank_account,
)
from apps.identity.services import (
    MemberError,
    assignable_users,
    company_members,
    member_work,
    # aliased: `add_member` already means "add someone to a work item"
    add_member as add_company_member,
    selectable_roles,
    set_member_role,
    set_member_status,
    set_password,
)
from apps.finance.models import Invoice
from apps.finance.services import (
    budget_vs_actual,
    commercial_dashboard,
    create_progress_claim,
    profit_forecast,
    profitability,
    rebuild_actuals_from_sources,
    record_payment,
)
from apps.procurement.models import GRN, GRNLine, PurchaseOrder, Supplier
from apps.procurement.services import three_way_match
from apps.projects.models import Project, ProjectStatus
from apps.quotes.models import Quotation, QuotationStatus
from apps.quotes.pdf import quotation_pdf_bytes
from apps.rfq.models import RFQDocument
from apps.rfq.models import RFQLineItem
from apps.rfq.services import (
    add_line_item,
    approve_rfq,
    delete_line_item,
    ingest_rfq,
    ingest_rfq_text,
    update_line_item,
)


def login_view(request):
    if request.user.is_authenticated:
        return redirect("web:dashboard")
    if request.method == "POST":
        user = authenticate(
            request,
            username=request.POST.get("email", "").strip(),
            password=request.POST.get("password", ""),
        )
        if user is not None:
            login(request, user)
            return redirect("web:dashboard")
        messages.error(request, "Invalid email or password.")
    return render(request, "web/login.html")


def logout_view(request):
    logout(request)
    return redirect("web:login")


def _can_view_money(user) -> bool:
    return user.has_perm_code("finance.view_money")


_STATUS_META = {
    "draft": ("#c4c7d0", "Draft"), "ready": ("#579bfc", "Ready"),
    "assigned": ("#579bfc", "Assigned"), "accepted": ("#a25ddc", "Accepted"),
    "in_progress": ("#fdab3d", "In progress"), "waiting": ("#fdab3d", "Waiting"),
    "blocked": ("#e2445c", "Blocked"), "quality_check": ("#a25ddc", "Quality check"),
    "client_signoff": ("#a25ddc", "Client sign-off"), "completed": ("#00c875", "Completed"),
    "closed": ("#676879", "Closed"), "cancelled": ("#c4c7d0", "Cancelled"),
}
_ORIGIN_META = {
    "rfq": ("#a25ddc", "RFQ / Tender"), "manual": ("#579bfc", "Manual work"),
    "project": ("#17a2b8", "Project"),
    "customer_request": ("#fdab3d", "Customer request"),
    "recurring": ("#00c875", "Recurring maintenance"),
    "internal": ("#676879", "Internal work"),
    "breakdown": ("#e2445c", "Breakdown / callout"),
    "preventative": ("#0e7c8c", "Preventative maintenance"),
}


def _donut(status_counts, total):
    """Build SVG donut segments (dash/gap/rotate computed server-side — no JS)."""
    from math import pi
    circumference = 2 * pi * 54
    segments, cum = [], 0.0
    for status, cnt in sorted(status_counts.items(), key=lambda x: -x[1]):
        color, label = _STATUS_META.get(status, ("#17a2b8", status.replace("_", " ").title()))
        pct = cnt / total if total else 0
        segments.append({
            "color": color, "label": label, "count": cnt, "pct": round(pct * 100),
            "dash": round(pct * circumference, 2), "gap": round((1 - pct) * circumference, 2),
            "rotate": round(-90 + cum * 360, 2),
        })
        cum += pct
    return {"segments": segments, "circ": round(circumference, 2)}


@login_required
def dashboard(request):
    """Portfolio home: KPI widgets + charts, plus the compliance attention list.
    The commercial panel is finance-only (Golden Rule)."""
    from collections import Counter

    projects = list(Project.objects.all().select_related("quotation"))
    attention = []
    for p in projects:
        r = recompute_readiness(p)
        if r["gate_status"] == "not_ready":
            attention.append({"project": p, "readiness": r})

    work = list(Task.objects.all())
    status_counts = Counter(t.status for t in work)
    origin_counts = Counter(t.origin for t in work)
    max_origin = max(origin_counts.values(), default=1)
    origins = [
        {"label": lbl, "color": col, "count": origin_counts.get(key, 0),
         "pct": round(origin_counts.get(key, 0) / max_origin * 100) if max_origin else 0}
        for key, (col, lbl) in _ORIGIN_META.items() if origin_counts.get(key, 0)
    ]

    context = {
        "project_count": len(projects),
        "attention": attention,
        "in_execution": sum(1 for p in projects if p.status == ProjectStatus.IN_EXECUTION),
        "ready": sum(1 for p in projects if p.status == ProjectStatus.READY),
        "work_total": len(work),
        "work_completed": status_counts.get("completed", 0),
        "work_blocked": status_counts.get("blocked", 0),
        "donut": _donut(status_counts, len(work)),
        "origins": origins,
        "can_view_money": _can_view_money(request.user),
    }
    if context["can_view_money"]:
        context["commercial"] = commercial_dashboard(request.user.active_company)
    return render(request, "web/dashboard.html", context)


@login_required
def projects_list(request):
    projects = Project.objects.all().select_related("quotation")
    return render(request, "web/projects.html", {"projects": projects})


@login_required
def project_detail(request, pk):
    project = get_object_or_404(Project.objects.all(), pk=pk)
    readiness = recompute_readiness(project)
    phases = list(project.phases.prefetch_related("tasks").all())
    unphased = list(project.tasks.filter(phase__isnull=True))
    context = {
        "project": project,
        "readiness": readiness,
        "health": project_health(project, request.user),
        "checklist": project.compliance_items.all(),
        "phases": phases,
        "unphased_tasks": unphased,
        "has_tasks": bool(unphased) or any(p.tasks.all() for p in phases),
        "can_view_money": _can_view_money(request.user),
        "can_compliance": request.user.has_perm_code("compliance.override"),
        "can_finance": request.user.has_perm_code("finance.manage"),
        "today": timezone.localdate().isoformat(),
    }
    if context["can_view_money"]:
        rebuild_actuals_from_sources(project, request.user)
        context["profitability"] = profitability(project)
        context["forecast"] = profit_forecast(project)
        context["budget"] = budget_vs_actual(project)
    return render(request, "web/project_detail.html", context)


@login_required
def readiness_partial(request, pk):
    """HTMX partial — re-renders just the readiness gate card, live."""
    project = get_object_or_404(Project.objects.all(), pk=pk)
    return render(request, "web/_readiness.html",
                  {"project": project, "readiness": recompute_readiness(project)})


# ══════════════════════════════════════════════════════════════════════════════
# Work Management Engine (MODULE 8) — one engine, many views
# ══════════════════════════════════════════════════════════════════════════════

#: Board columns — the lifecycle collapsed into the lanes a manager works in.
BOARD_LANES = [
    ("draft", "Draft", [TaskStatus.DRAFT]),
    ("ready", "Ready", [TaskStatus.READY, TaskStatus.ASSIGNED, TaskStatus.ACCEPTED]),
    ("active", "In progress", [TaskStatus.IN_PROGRESS]),
    ("stuck", "Stuck", [TaskStatus.BLOCKED, TaskStatus.WAITING]),
    ("review", "Review", [TaskStatus.QUALITY_CHECK, TaskStatus.CLIENT_SIGNOFF]),
    ("done", "Done", [TaskStatus.COMPLETED, TaskStatus.CLOSED]),
]

STATUS_TONE = {
    TaskStatus.COMPLETED: "ok", TaskStatus.CLOSED: "ok",
    TaskStatus.BLOCKED: "bad", TaskStatus.WAITING: "warn",
    TaskStatus.IN_PROGRESS: "warn",
    TaskStatus.QUALITY_CHECK: "purple", TaskStatus.CLIENT_SIGNOFF: "purple",
}

PRIORITY_TONE = {"critical": "bad", "high": "warn", "medium": "info",
                 "low": "", "planning": ""}


def _filtered_work(request):
    """One filter pipeline feeding every view — changing the view must never
    change the data."""
    qs = (Task.objects.all()
          .select_related("project", "phase", "assignee", "workspace")
          .prefetch_related("assignments__user"))
    f = request.GET
    if f.get("origin"):
        qs = qs.filter(origin=f["origin"])
    if f.get("status"):
        qs = qs.filter(status=f["status"])
    if f.get("priority"):
        qs = qs.filter(priority=f["priority"])
    if f.get("project"):
        qs = qs.filter(project_id=f["project"])
    if f.get("q"):
        qs = qs.filter(name__icontains=f["q"])
    scope = f.get("scope")
    if scope == "standalone":
        qs = qs.filter(project__isnull=True)
    elif scope == "project":
        qs = qs.filter(project__isnull=False)
    if f.get("mine"):
        qs = qs.filter(assignments__user=request.user).distinct()
    return qs


@login_required
def work_list(request):
    """Every unit of work, whatever its origin — one dataset, six lenses.
    The `view` parameter changes presentation only; filters are shared."""
    view = request.GET.get("view", "list")
    tasks = list(_filtered_work(request))

    ctx = {
        "tasks": tasks, "view": view,
        "origins": WorkOrigin.choices, "statuses": TaskStatus.choices,
        "priorities": TaskPriority.choices,
        "projects": Project.objects.all(),
        "f": request.GET,
        "status_tone": STATUS_TONE, "priority_tone": PRIORITY_TONE,
        "can_manage": has_work_perm(request.user, "work.edit"),
        "report": portfolio_report(),
    }

    if view == "board":
        ctx["lanes"] = [
            {"key": key, "label": label,
             "tasks": [t for t in tasks if t.status in members]}
            for key, label, members in BOARD_LANES
        ]
    elif view == "calendar":
        buckets = {}
        for t in tasks:
            if t.due_date:
                buckets.setdefault(t.due_date, []).append(t)
        ctx["calendar"] = sorted(buckets.items())
        ctx["undated"] = [t for t in tasks if not t.due_date]
    elif view == "workload":
        load = {}
        for t in tasks:
            for a in t.assignments.all():
                if a.role == Assignment.Role.WATCHER:
                    continue
                row = load.setdefault(a.user, {"user": a.user, "open": 0, "overdue": 0,
                                               "hours": Decimal("0")})
                if t.is_open:
                    row["open"] += 1
                    row["hours"] += t.estimated_hours or Decimal("0")
                if t.is_overdue:
                    row["overdue"] += 1
        rows = sorted(load.values(), key=lambda r: -r["open"])
        peak = max([r["open"] for r in rows], default=0) or 1
        for r in rows:
            r["pct"] = round(100 * r["open"] / peak)
        ctx["workload"] = rows
        ctx["unassigned"] = [t for t in tasks if t.is_open and not t.assignments.all()]

    return render(request, "web/work.html", ctx)


@login_required
def work_new(request):
    """The universal New Work wizard — the same front door for every origin."""
    if not has_work_perm(request.user, "work.create"):
        messages.error(request, "You do not have permission to create work.")
        return redirect("web:work")

    company = request.user.active_company
    if request.method == "POST":
        project = None
        if request.POST.get("project"):
            project = get_object_or_404(Project.objects.all(), pk=request.POST["project"])
        phase = None
        if request.POST.get("phase"):
            phase = get_object_or_404(Phase.objects.all(), pk=request.POST["phase"])

        owner = _user_or_none(request.POST.get("owner"))
        executors = [_user_or_none(u) for u in request.POST.getlist("executors")]
        watchers = [_user_or_none(u) for u in request.POST.getlist("watchers")]
        approvers = [_user_or_none(u) for u in request.POST.getlist("approvers")]
        labels = [s.strip() for s in request.POST.get("labels", "").split(",") if s.strip()]

        task = create_work(
            company, request.user,
            name=request.POST.get("name", "").strip() or "Untitled work",
            description=request.POST.get("description", "").strip(),
            origin=request.POST.get("origin") or WorkOrigin.MANUAL,
            project=project, phase=phase,
            priority=request.POST.get("priority") or TaskPriority.MEDIUM,
            risk_level=request.POST.get("risk_level") or RiskLevel.LOW,
            site=request.POST.get("site", "").strip(),
            department=request.POST.get("department", "").strip(),
            due_date=request.POST.get("due_date") or None,
            estimated_hours=_decimal_or_none(request.POST.get("estimated_hours")),
            labels=labels,
            is_billable=bool(request.POST.get("is_billable")),
            client_name=request.POST.get("client_name", "").strip(),
            owner=owner or request.user,
            executors=executors, watchers=watchers, approvers=approvers,
        )
        for line in request.POST.get("checklist", "").splitlines():
            if line.strip():
                add_checklist_item(task, request.user, label=line.strip())
        for f in request.FILES.getlist("attachments"):
            add_attachment(task, request.user, uploaded_file=f)

        messages.success(request, f"Work created: {task.name}.")
        return redirect("web:work_detail", pk=task.id)

    return render(request, "web/work_new.html", {
        "origins": WorkOrigin.choices,
        "priorities": TaskPriority.choices,
        "risks": RiskLevel.choices,
        "projects": Project.objects.all(),
        "people": _company_users(request.user),
    })


@login_required
def work_detail(request, pk):
    """The work item's own dashboard — progress, blockers, team, hierarchy,
    conversation and files in one place."""
    task = get_object_or_404(
        Task.objects.all().select_related("project", "phase", "assignee"), pk=pk)
    ctx = work_dashboard(task, request.user)
    ctx.update({
        "can_manage": has_work_perm(request.user, "work.edit"),
        "people": _company_users(request.user),
        "roles": Assignment.Role.choices,
        "dep_kinds": TaskDependency.Kind.choices,
        "status_tone": STATUS_TONE, "priority_tone": PRIORITY_TONE,
        "sibling_tasks": Task.objects.all().exclude(pk=task.pk)[:100],
        "subtasks": task.subtasks.prefetch_related("checklist_items"),
        "loose_checklist": task.checklist_items.filter(subtask__isnull=True),
    })
    return render(request, "web/work_detail.html", ctx)


@login_required
def work_operations(request, pk):
    """The manager's read-out of everything the field captured on a task: money
    allocated vs spent, materials, the GPS check-ins (plotted against the
    expected site), documents and the operational timeline."""
    import math

    from apps.execution.work_execution import task_operational_dashboard

    task = get_object_or_404(
        Task.objects.all().select_related("project", "phase", "assignee"), pk=pk)
    data = task_operational_dashboard(task, request.user)

    # Self-contained scatter map: each located check-in as a metre-offset from the
    # expected site, scaled into a 300×300 SVG so a manager sees at a glance which
    # reports fell outside the tolerance ring.
    map_ctx = None
    pts = data["map_points"]
    if task.site_latitude is not None and task.site_longitude is not None and pts:
        slat, slon = float(task.site_latitude), float(task.site_longitude)
        mlat = 111_320.0
        mlon = 111_320.0 * math.cos(math.radians(slat))
        offs = [{
            "dx": (p["lng"] - slon) * mlon,
            "dy": (p["lat"] - slat) * mlat,
            "flagged": p["flagged"], "title": p["title"],
        } for p in pts]
        tol = float(task.project.gps_tolerance_m if task.project_id else 500)
        span = max([tol] + [math.hypot(o["dx"], o["dy"]) for o in offs]) or 1.0
        scale = 130.0 / span  # px per metre, leaving a margin in the 300 box
        map_ctx = {
            "ring_px": tol * scale,
            "points": [{
                "x": 150 + o["dx"] * scale,
                "y": 150 - o["dy"] * scale,  # SVG y grows downward; north is up
                "flagged": o["flagged"], "title": o["title"],
            } for o in offs],
        }

    return render(request, "web/work_operations.html", {
        **data, "task": task, "map": map_ctx,
        "can_manage": has_work_perm(request.user, "work.edit"),
    })


def _work_guard(request, pk, code="work.edit"):
    """Shared permission gate for the work action endpoints. Each action names
    the granular permission it needs; `execution.manage` covers them all."""
    task = get_object_or_404(Task.objects.all(), pk=pk)
    if not has_work_perm(request.user, code):
        messages.error(request, "You do not have permission.")
        return task, False
    return task, True


@login_required
@require_POST
def work_start(request, pk):
    task, allowed = _work_guard(request, pk)
    if allowed:
        try:
            start_task(task, request.user)
        except ValueError as exc:
            messages.error(request, str(exc))
        else:
            messages.success(request, f"Started: {task.name}.")
    return redirect("web:work_detail", pk=pk)


@login_required
@require_POST
def work_complete(request, pk):
    task, allowed = _work_guard(request, pk)
    if allowed:
        complete_task(task, request.user,
                      actual_hours=request.POST.get("actual_hours") or None)
        messages.success(request, f"Completed: {task.name}.")
    return redirect("web:work_detail", pk=pk)


#: Lifecycle moves that are more than an edit — they need their own permission.
_TRANSITION_PERM = {
    TaskStatus.QUALITY_CHECK: "work.approve",
    TaskStatus.CLIENT_SIGNOFF: "work.approve",
    TaskStatus.CLOSED: "work.close",
}


@login_required
@require_POST
def work_transition(request, pk):
    """Move the work through the lifecycle. Sign-off and closure are held to a
    higher permission than an ordinary edit."""
    target = request.POST.get("status")
    task, allowed = _work_guard(request, pk, _TRANSITION_PERM.get(target, "work.edit"))
    if allowed:
        if target not in dict(TaskStatus.choices):
            messages.error(request, "Unknown status.")
        else:
            try:
                transition(task, request.user, to_status=target,
                           note=request.POST.get("note", ""))
            except ValueError as exc:
                messages.error(request, str(exc))
            else:
                messages.success(request, f"{task.name} → {task.get_status_display()}.")
    return redirect("web:work_detail", pk=pk)


@login_required
@require_POST
def work_subtask_add(request, pk):
    task, allowed = _work_guard(request, pk)
    if allowed and request.POST.get("name", "").strip():
        add_subtask(task, request.user, name=request.POST["name"].strip(),
                    assignee=_user_or_none(request.POST.get("assignee")),
                    due_date=request.POST.get("due_date") or None)
    return redirect("web:work_detail", pk=pk)


@login_required
@require_POST
def work_checklist_add(request, pk):
    task, allowed = _work_guard(request, pk)
    if allowed and request.POST.get("label", "").strip():
        subtask = None
        if request.POST.get("subtask"):
            subtask = get_object_or_404(Subtask.objects.all(), pk=request.POST["subtask"])
        add_checklist_item(task, request.user, label=request.POST["label"].strip(),
                           subtask=subtask)
    return redirect("web:work_detail", pk=pk)


@login_required
@require_POST
def work_checklist_toggle(request, pk, item_id):
    """Ticking an item rolls progress up the hierarchy — anyone on the work may
    do it (that is the point of a checklist), watchers excepted."""
    task = get_object_or_404(Task.objects.all(), pk=pk)
    item = get_object_or_404(ChecklistItem.objects.filter(task=task), pk=item_id)
    from apps.execution.services import can_modify
    if can_modify(task, request.user):
        toggle_checklist_item(item, request.user)
    else:
        messages.error(request, "You do not have permission.")
    return redirect("web:work_detail", pk=pk)


@login_required
@require_POST
def work_comment_add(request, pk):
    task = get_object_or_404(Task.objects.all(), pk=pk)
    body = request.POST.get("body", "").strip()
    if body:
        parent = None
        if request.POST.get("parent"):
            parent = task.comments.filter(pk=request.POST["parent"]).first()
        add_comment(task, request.user, body=body, parent=parent,
                    is_internal=not request.POST.get("customer_visible"))
    for f in request.FILES.getlist("files"):
        add_attachment(task, request.user, uploaded_file=f)
    return redirect("web:work_detail", pk=pk)


@login_required
@require_POST
def work_file_add(request, pk):
    task, allowed = _work_guard(request, pk, "work.files")
    if allowed:
        for f in request.FILES.getlist("files"):
            add_attachment(task, request.user, uploaded_file=f,
                           kind=request.POST.get("kind", "document"))
    return redirect("web:work_detail", pk=pk)


@login_required
@require_POST
def work_member(request, pk):
    task, allowed = _work_guard(request, pk, "work.assign")
    if allowed:
        user = _user_or_none(request.POST.get("user"))
        role = request.POST.get("role")
        if user and role in dict(Assignment.Role.choices):
            if request.POST.get("remove"):
                remove_member(task, user, role)
            else:
                add_member(task, user, role)
    return redirect("web:work_detail", pk=pk)


@login_required
@require_POST
def work_link(request, pk):
    task, allowed = _work_guard(request, pk)
    if allowed:
        other = Task.objects.filter(pk=request.POST.get("from_task")).first()
        if other:
            try:
                link_tasks(other, task, kind=request.POST.get("kind") or None)
            except ValueError as exc:
                messages.error(request, str(exc))
            else:
                messages.success(request, "Dependency added.")
    return redirect("web:work_detail", pk=pk)


@login_required
def work_decompose(request, pk):
    """LulaAI proposes a breakdown for this work. Read-only: computing a proposal
    writes nothing, so a manager can ask as often as they like and only what they
    tick ever becomes real."""
    task = get_object_or_404(Task.objects.all(), pk=pk)
    if not request.user.has_perm_code("ai.generate"):
        messages.error(request, "You do not have permission to use LulaAI.")
        return redirect("web:work_detail", pk=pk)

    draft = propose_decomposition(
        request.user.active_company, request.user,
        name=task.name, description=task.description,
        origin=task.origin, project=task.project,
    )
    record_proposal(request.user.active_company, request.user, task, draft)
    return render(request, "web/work_decompose.html", {
        "task": task, "draft": draft,
        "existing_checklist": {c.label.lower() for c in task.checklist_items.all()},
        "can_manage": has_work_perm(request.user, "work.edit"),
    })


@login_required
@require_POST
def work_decompose_apply(request, pk):
    """Create ONLY the ticked items. The draft is recomputed (it is deterministic)
    and the posted indexes select from it — nothing is trusted from the client
    except which items the human approved."""
    task, allowed = _work_guard(request, pk)
    if not allowed:
        return redirect("web:work_detail", pk=pk)
    if not request.user.has_perm_code("ai.generate"):
        messages.error(request, "You do not have permission to use LulaAI.")
        return redirect("web:work_detail", pk=pk)

    draft = propose_decomposition(
        request.user.active_company, request.user,
        name=task.name, description=task.description,
        origin=task.origin, project=task.project, enrich=False,
    )
    checklist_indexes = {int(i) for i in request.POST.getlist("checklist") if i.isdigit()}
    phase_indexes = {int(i) for i in request.POST.getlist("phases") if i.isdigit()}
    applied = apply_decomposition(
        task, request.user, draft,
        checklist_indexes=checklist_indexes, phase_indexes=phase_indexes,
        apply_hours=bool(request.POST.get("apply_hours")),
    )
    record_proposal(request.user.active_company, request.user, task, draft, applied=applied)

    if applied["checklist"] or applied["phases"] or applied["hours_set"]:
        messages.success(
            request,
            f"Added {applied['checklist']} checklist item(s)"
            f"{', ' + str(applied['phases']) + ' phase(s)' if applied['phases'] else ''}"
            f"{', set the estimate' if applied['hours_set'] else ''}.")
    else:
        messages.error(request, "Nothing selected — nothing was created.")
    return redirect("web:work_detail", pk=pk)


@login_required
def notifications(request):
    from apps.execution.models import Notification
    rows = Notification.objects.filter(user=request.user).select_related("task")[:100]
    if request.method == "POST":
        mark_notifications_read(request.user)
        return redirect("web:notifications")
    return render(request, "web/notifications.html",
                  {"rows": rows, "unread": unread_count(request.user)})


@login_required
@require_POST
def project_phase_add(request, pk):
    project = get_object_or_404(Project.objects.all(), pk=pk)
    if not has_work_perm(request.user, "work.edit"):
        messages.error(request, "You do not have permission.")
    elif request.POST.get("seed"):
        ensure_default_phases(project, request.user)
        messages.success(request, "Default phases added.")
    elif request.POST.get("name", "").strip():
        add_phase(project, request.user, name=request.POST["name"].strip())
        messages.success(request, "Phase added.")
    return redirect("web:project_detail", pk=pk)


# ══════════════════════════════════════════════════════════════════════════════
# Company profile — the identity every other module reads from
# ══════════════════════════════════════════════════════════════════════════════

#: Which POSTed fields each section owns. Saving a section touches only its own
#: fields, so two people editing different sections cannot clobber each other.
_PROFILE_SECTIONS = {
    "identity": ["name", "trading_name", "registration_no", "tax_reference_no",
                 "vat_no", "company_type", "industry", "year_established"],
    "contact": ["email", "phone", "phone_secondary", "mobile", "whatsapp",
                "emergency_phone", "website", "facebook", "linkedin", "twitter"],
    "address": ["country", "province", "city", "suburb", "street_address",
                "postal_code"],
    "postal": ["postal_same_as_physical", "postal_address", "postal_city",
               "postal_code_postal", "postal_country"],
    "business": ["description", "employee_count", "vehicle_count", "site_count"],
}

#: Comma-separated inputs stored as lists.
_LIST_FIELDS = {"business": ["services_offered", "specialisations",
                             "industries_served", "operating_provinces",
                             "operating_countries"]}
_INT_FIELDS = {"year_established", "employee_count", "vehicle_count", "site_count"}


@login_required
def company_profile(request):
    """One page, many sections. Everything here is read by quotations, invoices,
    purchase orders and every generated PDF — it is entered once, here."""
    company = get_profile(request.user.active_company)
    can_edit = request.user.has_perm_code("company.manage")

    if request.method == "POST":
        if not can_edit:
            messages.error(request, "You do not have permission to edit the company profile.")
            return redirect("web:company_profile")
        section = request.POST.get("section", "")
        if section in _PROFILE_SECTIONS:
            _save_profile_section(request, company, section)
            messages.success(request, "Company profile updated.")
        elif section == "compliance":
            _save_compliance(request, company)
            messages.success(request, "Statutory details updated.")
        elif section == "branding":
            _save_branding(request, company)
            messages.success(request, "Branding updated.")
        elif section == "defaults":
            _save_defaults(request, company)
            messages.success(request, "Defaults updated.")
        elif section == "commercial":
            _save_commercial(request, company)
            messages.success(request, "Commercial document settings updated.")
        else:
            messages.error(request, "Unknown section.")
        return redirect("web:company_profile")

    from apps.administration.models import CompanySettings
    return render(request, "web/company.html", {
        "company": company,
        "compliance": company.compliance,
        "branding": company.branding,
        "settings": CompanySettings.objects.get(company=company),
        "bank_accounts": company.bank_accounts.all(),
        "contacts": company.contacts.all(),
        "documents": company.documents.all(),
        "score": completeness(company),
        "default_bank": default_bank_account(company),
        "expiring": company.compliance.expiring(),
        "can_edit": can_edit,
    })


def _save_profile_section(request, company, section):
    for field in _PROFILE_SECTIONS[section]:
        if field not in request.POST and field != "postal_same_as_physical":
            continue
        value = request.POST.get(field, "").strip()
        if field == "postal_same_as_physical":
            setattr(company, field, bool(request.POST.get(field)))
        elif field in _INT_FIELDS:
            setattr(company, field, int(value) if value.isdigit() else None)
        else:
            setattr(company, field, value)
    for field in _LIST_FIELDS.get(section, []):
        raw = request.POST.get(field, "")
        setattr(company, field, [v.strip() for v in raw.split(",") if v.strip()])
    company.save()


def _save_compliance(request, company):
    c = company.compliance
    c.vat_registered = bool(request.POST.get("vat_registered"))
    for field in ("income_tax_no", "paye_no", "uif_no", "coida_no",
                  "bbbee_level", "csd_supplier_no", "cidb_grading"):
        setattr(c, field, request.POST.get(field, "").strip())
    for field in ("coida_expiry", "bbbee_expiry"):
        setattr(c, field, request.POST.get(field) or None)
    for field in ("iso_certifications", "industry_certifications"):
        raw = request.POST.get(field, "")
        setattr(c, field, [v.strip() for v in raw.split(",") if v.strip()])
    c.save()


def _clean_hex_colour(raw):
    """A valid #RRGGBB / #RRGGBBAA hex, '' to clear, or None if the value was
    given but is not a colour — so a mistyped colour is skipped with a warning
    rather than crashing the page (the field is only 9 characters wide)."""
    s = (raw or "").strip()
    if not s:
        return ""
    if not s.startswith("#"):
        s = "#" + s
    if re.fullmatch(r"#[0-9a-fA-F]{3}", s):          # #abc → #aabbcc
        s = "#" + "".join(c * 2 for c in s[1:])
    return s.lower() if re.fullmatch(r"#(?:[0-9a-fA-F]{6}|[0-9a-fA-F]{8})", s) else None


def _save_branding(request, company):
    if request.FILES.get("logo"):
        company.logo = request.FILES["logo"]
        company.save(update_fields=["logo"])
    branding = company.branding
    changed = []
    for slot in ("email_logo", "invoice_logo", "report_logo", "letterhead",
                 "stamp", "signature", "seal"):
        if request.FILES.get(slot):
            setattr(branding, slot, request.FILES[slot])
            changed.append(slot)
    for colour in ("brand_primary", "brand_secondary"):
        if colour in request.POST:
            cleaned = _clean_hex_colour(request.POST.get(colour, ""))
            if cleaned is None:
                messages.warning(request, "Brand colour must be a hex value like "
                                          "#0E6E6E — that entry was ignored.")
            else:
                setattr(company, colour, cleaned)
    # The document prefix — letters only, uppercased, at most four. Begins every
    # commercial reference (quotation, and from it invoices and delivery notes).
    saved_fields = ["brand_primary", "brand_secondary"]
    if "document_prefix" in request.POST:
        prefix = "".join(c for c in request.POST.get("document_prefix", "")
                         if c.isalpha()).upper()[:4]
        company.document_prefix = prefix
        saved_fields.append("document_prefix")
    company.save(update_fields=saved_fields)
    if changed:
        branding.save(update_fields=changed)


def _save_defaults(request, company):
    from apps.administration.models import CompanySettings
    settings_row = CompanySettings.objects.get(company=company)
    for field in ("date_format", "number_format", "language", "ai_provider",
                  "ai_language", "ai_response_style"):
        if field in request.POST:
            setattr(settings_row, field, request.POST.get(field, "").strip())
    for field in ("financial_year_start_month", "week_starts_on"):
        value = request.POST.get(field, "")
        if value.isdigit():
            setattr(settings_row, field, int(value))
    if request.POST.get("tax_rate"):
        settings_row.tax_rate = _decimal_or_none(request.POST["tax_rate"]) or settings_row.tax_rate
    for flag in ("ai_suggestions_enabled", "ai_summaries_enabled",
                 "ai_cost_estimation_enabled", "ai_task_generation_enabled",
                 "ai_compliance_detection_enabled"):
        setattr(settings_row, flag, bool(request.POST.get(flag)))
    settings_row.save()
    for field in ("currency", "timezone"):
        if field in request.POST:
            setattr(company, field, request.POST.get(field, "").strip())
    company.save(update_fields=["currency", "timezone"])


def _save_commercial(request, company):
    """Standard terms & conditions per document type — configured once, then
    auto-inserted into every quotation, invoice and delivery note."""
    from apps.administration.models import CompanySettings
    settings_row = CompanySettings.objects.get(company=company)
    for field in ("quotation_terms", "invoice_terms", "delivery_terms"):
        if field in request.POST:
            setattr(settings_row, field, request.POST.get(field, "").strip())
    settings_row.save(update_fields=["quotation_terms", "invoice_terms",
                                     "delivery_terms", "updated_at"])


@login_required
@require_POST
def company_hours(request):
    """The company calendar: the week, lunch, emergency cover, and holidays.

    This is not decoration — `add_working_days` reads it, so a change here
    changes what "due in 5 days" means everywhere.
    """
    from apps.administration.hours import (
        DAYS,
        add_holiday,
        ensure_statutory_holidays,
        remove_holiday,
        save_week,
    )
    company = request.user.active_company
    if not request.user.has_perm_code("company.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:company_hours_page")

    action = request.POST.get("action", "week")
    if action == "week":
        week = {}
        for key in DAYS:
            week[key] = {
                "closed": not request.POST.get(f"{key}_enabled"),
                "open": request.POST.get(f"{key}_open", "").strip(),
                "close": request.POST.get(f"{key}_close", "").strip(),
                "lunch_start": request.POST.get(f"{key}_lunch_start", "").strip(),
                "lunch_end": request.POST.get(f"{key}_lunch_end", "").strip(),
            }
        save_week(company, week)

        from apps.administration.models import CompanySettings
        row = CompanySettings.objects.get(company=company)
        row.emergency_support = bool(request.POST.get("emergency_support"))
        row.emergency_hours = request.POST.get("emergency_hours", "").strip()
        row.emergency_note = request.POST.get("emergency_note", "").strip()
        row.save(update_fields=["emergency_support", "emergency_hours", "emergency_note"])
        messages.success(request, "Working hours updated.")

    elif action == "seed_holidays":
        year = int(request.POST.get("year") or timezone.localdate().year)
        added = ensure_statutory_holidays(company, year)
        messages.success(
            request,
            f"Added {added} public holiday(s) for {year}." if added
            else f"{year} already has every statutory holiday.")

    elif action == "add_holiday":
        day, name = request.POST.get("date"), request.POST.get("name", "").strip()
        if not day or not name:
            messages.error(request, "A date and a name are required.")
        else:
            add_holiday(company, day=day, name=name)
            messages.success(request, "Closure added.")

    elif action == "remove_holiday":
        remove_holiday(company, request.POST.get("date"))
        messages.success(request, "Removed.")

    return redirect("web:company_hours_page")


@login_required
def company_hours_page(request):
    from apps.administration.hours import (
        DAY_LABELS,
        DAYS,
        get_week,
        holidays,
        is_open,
        weekly_hours,
    )
    from apps.administration.models import CompanySettings
    from apps.execution.services import due_date_from_duration

    company = request.user.active_company
    week = get_week(company)
    today = timezone.localdate()

    return render(request, "web/company_hours.html", {
        "company": company,
        "settings": CompanySettings.objects.get_or_create(company=company)[0],
        "days": [{"key": k, "label": DAY_LABELS[k], **week[k]} for k in DAYS],
        "status": is_open(company),
        "weekly_hours": weekly_hours(company),
        "holidays": holidays(company, today.year),
        "next_year": today.year + 1,
        "this_year": today.year,
        # Proof the calendar is load-bearing rather than decorative.
        "example": {
            "days": 5,
            "calendar": today + timezone.timedelta(days=5),
            "working": due_date_from_duration(company, 5, today),
        },
        "can_edit": request.user.has_perm_code("company.manage"),
    })


@login_required
@require_POST
def company_bank(request):
    """Add, default, or remove a bank account. These print on invoices, so the
    action is deliberately explicit rather than inline-editable."""
    company = request.user.active_company
    if not request.user.has_perm_code("company.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:company_profile")

    action = request.POST.get("action", "add")
    if action == "add":
        if not request.POST.get("bank_name") or not request.POST.get("account_number"):
            messages.error(request, "Bank name and account number are required.")
        else:
            add_bank_account(
                company,
                bank_name=request.POST["bank_name"].strip(),
                account_name=request.POST.get("account_name", "").strip() or company.name,
                account_number=request.POST["account_number"].strip(),
                branch_name=request.POST.get("branch_name", "").strip(),
                branch_code=request.POST.get("branch_code", "").strip(),
                account_type=request.POST.get("account_type", "cheque"),
                swift_code=request.POST.get("swift_code", "").strip(),
                currency=request.POST.get("currency", "").strip() or company.currency,
                label=request.POST.get("label", "").strip(),
            )
            messages.success(request, "Bank account added.")
    else:
        account = get_object_or_404(
            CompanyBankAccount.objects.filter(company=company),
            pk=request.POST.get("account"))
        if action == "default":
            set_default_bank_account(account)
            messages.success(request, f"{account.bank_name} is now the default account.")
        elif action == "delete":
            account.delete()
            messages.success(request, "Bank account removed.")
    return redirect("web:company_profile")


@login_required
@require_POST
def company_contact(request):
    company = request.user.active_company
    if not request.user.has_perm_code("company.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:company_profile")

    action = request.POST.get("action", "add")
    if action == "add":
        if not request.POST.get("full_name"):
            messages.error(request, "A name is required.")
        else:
            add_contact(
                company,
                full_name=request.POST["full_name"].strip(),
                job_title=request.POST.get("job_title", "").strip(),
                email=request.POST.get("email", "").strip(),
                phone=request.POST.get("phone", "").strip(),
                mobile=request.POST.get("mobile", "").strip(),
                extension=request.POST.get("extension", "").strip(),
                preferred_method=request.POST.get("preferred_method", "email"),
            )
            messages.success(request, "Contact added.")
    else:
        contact = get_object_or_404(
            CompanyContact.objects.filter(company=company), pk=request.POST.get("contact"))
        if action == "primary":
            contact.is_primary = True
            contact.save()
            messages.success(request, f"{contact.full_name} is now the primary contact.")
        elif action == "delete":
            contact.delete()
            messages.success(request, "Contact removed.")
    return redirect("web:company_profile")


@login_required
@require_POST
def company_document(request):
    company = request.user.active_company
    if not request.user.has_perm_code("company.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:company_profile")
    if request.POST.get("action") == "delete":
        doc = get_object_or_404(CompanyDocument.objects.filter(company=company),
                                pk=request.POST.get("document"))
        doc.delete()
        messages.success(request, "Document removed.")
    elif request.FILES.get("file"):
        CompanyDocument.objects.create(
            company=company, file=request.FILES["file"],
            name=request.POST.get("name", "").strip() or request.FILES["file"].name,
            doc_type=request.POST.get("doc_type", "").strip(),
            expires_on=request.POST.get("expires_on") or None,
        )
        messages.success(request, "Document uploaded.")
    else:
        messages.error(request, "Choose a file to upload.")
    return redirect("web:company_profile")


# ══════════════════════════════════════════════════════════════════════════════
# People — the company's own members (the pool every work picker reads from)
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def people(request):
    """Who works here. Managing members needs `users.invite`; everyone else may
    see the team (they work with these people)."""
    company = request.user.active_company
    members = company_members(company)
    return render(request, "web/people.html", {
        "members": members,
        "active_count": sum(1 for m in members if m.status == "active"),
        "roles": selectable_roles(),
        "can_manage": request.user.has_perm_code("users.invite"),
        # Surfaced once, immediately after creation — never stored or re-shown.
        "new_password": request.session.pop("new_member_password", None),
        "new_email": request.session.pop("new_member_email", None),
    })


@login_required
@require_POST
def people_add(request):
    if not request.user.has_perm_code("users.invite"):
        messages.error(request, "You do not have permission to manage members.")
        return redirect("web:people")

    role = Role.objects.filter(pk=request.POST.get("role")).first()
    if role is None:
        messages.error(request, "Choose a role for the new member.")
        return redirect("web:people")

    try:
        membership, temp_password = add_company_member(
            request.user.active_company, request.user,
            email=request.POST.get("email", ""),
            first_name=request.POST.get("first_name", ""),
            last_name=request.POST.get("last_name", ""),
            job_title=request.POST.get("job_title", ""),
            mobile=request.POST.get("mobile", ""),
            role=role,
        )
    except MemberError as exc:
        messages.error(request, str(exc))
        return redirect("web:people")

    if temp_password:
        # Carried in the session for exactly one render, then popped.
        request.session["new_member_password"] = temp_password
        request.session["new_member_email"] = membership.user.email
        messages.success(request, f"{membership.user.email} added.")
    else:
        messages.success(
            request,
            f"{membership.user.email} already had a LulaWorks account and has "
            "been added to this company with their existing password.")
    return redirect("web:people")


@login_required
@require_POST
def people_role(request, pk):
    if not request.user.has_perm_code("users.invite"):
        messages.error(request, "You do not have permission to manage members.")
        return redirect("web:people")
    membership = get_object_or_404(
        Membership.objects.filter(company=request.user.active_company), pk=pk)
    role = Role.objects.filter(pk=request.POST.get("role")).first()
    if role is None:
        messages.error(request, "Unknown role.")
    else:
        set_member_role(membership, role)
        messages.success(request, f"{membership.user.email} is now {role.name}.")
    return redirect("web:people")


@login_required
@require_POST
def people_status(request, pk):
    """Deactivate or restore. Never deletes — a departed employee stays attached
    to the work, timesheets and sign-offs they touched."""
    if not request.user.has_perm_code("users.invite"):
        messages.error(request, "You do not have permission to manage members.")
        return redirect("web:people")
    membership = get_object_or_404(
        Membership.objects.filter(company=request.user.active_company), pk=pk)
    activate = bool(request.POST.get("activate"))
    try:
        set_member_status(membership, request.user, active=activate)
    except MemberError as exc:
        messages.error(request, str(exc))
    else:
        messages.success(
            request,
            f"{membership.user.email} " + ("restored." if activate else "deactivated."))
    return redirect("web:people")


@login_required
def person_detail(request, pk):
    """One member: who they are, what they are on right now, and what they have
    finished. Reads through Assignment, so every role they hold shows up."""
    membership = get_object_or_404(
        Membership.objects.select_related("user", "role")
                  .filter(company=request.user.active_company), pk=pk)
    work = member_work(membership.user, request.user.active_company)
    return render(request, "web/person_detail.html", {
        "membership": membership,
        "person": membership.user,
        "work": work,
        "roles": selectable_roles(),
        "can_manage": request.user.has_perm_code("users.invite"),
        "is_me": membership.user_id == request.user.id,
    })


@login_required
def profile(request):
    """Your own profile — the one page every member can edit for themselves,
    whatever their role. Adding a photo is the first thing a new member does
    after signing in with the password they were handed."""
    user = request.user
    if request.method == "POST":
        if request.POST.get("remove_avatar") and user.avatar:
            user.avatar.delete(save=False)
            user.avatar = None
            user.save(update_fields=["avatar"])
            messages.success(request, "Photo removed.")
            return redirect("web:profile")

        upload = request.FILES.get("avatar")
        if upload is not None:
            error = _validate_avatar(upload)
            if error:
                messages.error(request, error)
                return redirect("web:profile")
            user.avatar = upload

        user.first_name = request.POST.get("first_name", "").strip()
        user.last_name = request.POST.get("last_name", "").strip()
        user.mobile = request.POST.get("mobile", "").strip()
        user.save(update_fields=["first_name", "last_name", "mobile", "avatar"])
        messages.success(request, "Profile updated.")
        return redirect("web:profile")

    membership = Membership.objects.filter(
        company=user.active_company, user=user).select_related("role").first()
    return render(request, "web/profile.html", {
        "person": user, "membership": membership,
        "work": member_work(user, user.active_company),
    })


#: Anything bigger is a phone photo nobody needs as an avatar.
MAX_AVATAR_BYTES = 5 * 1024 * 1024
ALLOWED_AVATAR_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}


def _validate_avatar(upload):
    """Reject by declared type AND by actually decoding the file — a renamed
    executable must not be stored just because it ends in .png."""
    if upload.size > MAX_AVATAR_BYTES:
        return "That image is larger than 5 MB — please use a smaller one."
    if upload.content_type not in ALLOWED_AVATAR_TYPES:
        return "Please upload a JPEG, PNG, WebP or GIF image."
    try:
        from PIL import Image
        image = Image.open(upload)
        image.verify()          # raises unless this really is an image
    except Exception:           # noqa: BLE001 - any decode failure is a rejection
        return "That file is not a readable image."
    finally:
        upload.seek(0)
    return None


@login_required
def change_password(request):
    """Choose your own password. Forced (via middleware) for accounts a manager
    created with a temporary one; available to anyone otherwise."""
    forced = request.user.must_change_password
    if request.method == "POST":
        current = request.POST.get("current_password", "")
        new = request.POST.get("new_password", "")
        confirm = request.POST.get("confirm_password", "")

        if not request.user.check_password(current):
            messages.error(request, "Your current password is not correct.")
        elif len(new) < 10:
            messages.error(request, "Choose at least 10 characters.")
        elif new != confirm:
            messages.error(request, "The two new passwords do not match.")
        elif new == current:
            messages.error(request, "The new password must be different.")
        else:
            set_password(request.user, new)
            # Changing the hash rotates the session auth hash — keep them signed in.
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, request.user)
            messages.success(request, "Password updated.")
            return redirect("web:dashboard")

    return render(request, "web/change_password.html", {"forced": forced})


# ── Small helpers for the work forms ──────────────────────────────────────────

def _company_users(user):
    """The pool every team picker offers: ACTIVE members of the acting user's
    company. Someone who has left stops being assignable but keeps their name on
    the work they already did."""
    return assignable_users(user.active_company)


def _user_or_none(value):
    if not value:
        return None
    from apps.identity.models import User
    return User.objects.filter(pk=value).first()


def _decimal_or_none(value):
    if not value:
        return None
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError):
        return None


def _positive_decimal(value):
    """A non-negative Decimal (0 when missing or negative) — for amounts a client
    must never be able to send negative, e.g. an overall discount."""
    d = _decimal_or_none(value) or Decimal("0")
    return d if d > 0 else Decimal("0")


def _save_quotation_uploads(request, quote, company):
    """Attach the uploaded documents to a quotation, keeping the valid ones and
    warning about any rejected by the upload policy (type / size)."""
    from apps.core.uploads import clean_uploads
    accepted, rejected = clean_uploads(request.FILES.getlist("documents"))
    for f in accepted:
        quote.documents.create(company=company, name=f.name,
                               doc_type="attachment", file=f)
    for reason in rejected:
        messages.warning(request, f"Skipped {reason}")


# ── Quotations (view · review · edit · download PDF) ──────────────────────────

@login_required
def quotations_list(request):
    """The commercial pipeline, not just a list of documents."""
    from apps.quotes.services import pipeline
    from django.core.paginator import Paginator

    # select_related the FKs the table renders per row (customer, quotation type,
    # preparer); prefetch the lines the total/margin roll up from — so a page of
    # rows is a couple of queries, not one per quotation.
    quotes = (Quotation.objects.all()
              .select_related("customer", "quotation_type", "prepared_by")
              .prefetch_related("lines", "customer_pos"))
    status = request.GET.get("status")
    if status:
        quotes = quotes.filter(status=status)
    if request.GET.get("q"):
        quotes = quotes.filter(number__icontains=request.GET["q"])
    if request.GET.get("customer"):
        quotes = quotes.filter(customer_id=request.GET["customer"])

    page = Paginator(quotes, 25).get_page(request.GET.get("page"))
    # Everything but the page cursor, so filters survive Previous/Next.
    carried = request.GET.copy()
    carried.pop("page", None)

    from apps.customers.models import Customer
    return render(request, "web/quotations.html", {
        "quotations": page,
        "carried": carried.urlencode(),
        "pipeline": pipeline(),
        "statuses": QuotationStatus.choices,
        "customers": Customer.objects.all(),
        "f": request.GET,
        "can_view_money": _can_view_money(request.user),
        "can_create": request.user.has_perm_code("quotes.create"),
    })


def _edit_url(pk) -> str:
    """The builder is the detail page in edit mode; every editing action returns
    here so the estimator keeps working rather than bouncing to the review."""
    from django.urls import reverse
    return reverse("web:quotation_detail", args=[pk]) + "?edit=1"


@login_required
def quotation_detail(request, pk):
    """Two faces of one URL. By default this is the review workspace — the
    quotation as the customer will see it, with the actions that move it through
    its commercial life. With ``?edit=1`` (and while the quotation is still
    editable) it is the builder. Creating and estimating happen in edit mode;
    reviewing, approving, finalizing and issuing happen in the default view.
    """
    quote = get_object_or_404(
        Quotation.objects.all().prefetch_related(
            "lines__section", "sections", "customer_pos", "events",
            "documents", "revisions"),
        pk=pk)

    editing = bool(request.GET.get("edit")) and quote.is_editable \
        and request.user.has_perm_code("quotes.create")
    if editing:
        return _quotation_build(request, quote)
    return _quotation_review(request, quote)


def _quotation_build(request, quote):
    """Edit mode: header, sections, priced lines, margin, the estimating helpers.
    Reachable only while the quotation is editable (the detail view guards it)."""
    from apps.quotes.estimating_ai import pricing_review
    from apps.quotes.models import LineCategory, QuotationType, VatMode

    grouped = []
    for section in quote.sections.all():
        rows = [line for line in quote.lines.all() if line.section_id == section.id]
        grouped.append({"section": section, "lines": rows})
    ungrouped = [line for line in quote.lines.all() if line.section_id is None]

    from apps.customers.models import Customer

    return render(request, "web/quotation_build.html", {
        "quote": quote,
        "grouped": grouped,
        "ungrouped": ungrouped,
        "categories": LineCategory.choices,
        "vat_modes": VatMode.choices,
        "types": QuotationType.objects.all(),
        "customers": Customer.objects.all(),
        "can_view_money": _can_view_money(request.user),
        # Deterministic and free, so the estimator sees it while building.
        "review": pricing_review(quote),
    })


def _commercial_timeline(quote):
    """The commercial life of the quotation, each stage with its date and the
    person responsible — read from the quotation's own event history and its
    linked documents. Invoice/delivery/payment stages read as pending until
    those documents exist."""
    events = list(quote.events.all())

    def when(status):
        e = next((e for e in events if e.to_status == status), None)
        return (e.created_at, e.actor) if e else (None, None)

    rows = [{"label": "Quotation created", "done": True,
             "date": quote.created_at, "user": quote.created_by}]
    # Approval is the final step; the customer's answer follows. (No separate
    # finalize/sent stages.)
    dt, user = when("approved")
    rows.append({"label": "Approved", "done": bool(dt), "date": dt, "user": user})

    po = quote.customer_pos.all().first()
    rows.append({"label": "Purchase order received", "done": bool(po),
                 "date": (po.po_date or po.created_at) if po else None,
                 "user": po.created_by if po else None})

    docs = list(quote.commercial_documents.all())
    inv = next((d for d in docs if d.kind == "invoice"), None)
    dn = next((d for d in docs if d.kind == "delivery"), None)
    rows.append({"label": "Tax invoice created", "done": bool(inv),
                 "date": inv.created_at if inv else None,
                 "user": inv.created_by if inv else None})
    rows.append({"label": "Delivery note created", "done": bool(dn),
                 "date": dn.created_at if dn else None,
                 "user": dn.created_by if dn else None})
    rows.append({"label": "Payment received", "done": False, "date": None, "user": None})
    return rows


def _quotation_review(request, quote):
    """Review mode: the document as the customer will see it, plus the actions
    that move it through approval, finalize, issue and award — and the single
    Purchase Orders section once it is out with the customer."""
    from apps.identity.profile import document_header
    from apps.quotes.services import (
        can_generate_documents,
        matching_purchase_order,
        traceability,
    )

    can_quote = request.user.has_perm_code("quotes.create")
    can_approve = request.user.has_perm_code("quotes.approve")
    can_download = request.user.has_perm_code("quotes.download")
    matched_po = matching_purchase_order(quote)
    context = {
        "quote": quote,
        "header": document_header(quote.company, kind="quotation"),
        "can_view_money": _can_view_money(request.user),
        # Editing a draft needs create; approving is a separate, authorised step.
        "can_edit": quote.is_editable and can_quote,
        # Approve is the single, final step — it locks the quotation and turns on
        # the outputs (PDF, Excel, Create invoice/delivery).
        "can_approve": can_approve and quote.status in (
            "draft", "review", "manager_approval", "commercial_approval"),
        "can_download": quote.is_finalized and can_download,
        "can_revise": can_quote and quote.is_finalized,
        "can_award": request.user.has_perm_code("projects.create"),
        # The optional PO section opens once the quotation is finalised (the
        # customer may send one back, but it is not required to invoice).
        "po_active": quote.is_finalized,
        "purchase_orders": list(quote.customer_pos.all()),
        # Invoice / delivery note may be raised once a price-matching PO is linked.
        "can_generate_docs": can_quote and can_generate_documents(quote),
        "matched_po": matched_po,
        "commercial_documents": list(quote.commercial_documents.all()),
        # Once a document exists, its button becomes "View …" instead of "Create".
        "existing_invoice": quote.commercial_documents.filter(kind="invoice").first(),
        "existing_delivery": quote.commercial_documents.filter(kind="delivery").first(),
        # Start Work (approved) → operational project; once it exists, Open Work.
        "can_start_work": request.user.has_perm_code("projects.create"),
        "existing_project": quote.projects.first(),
        "revisions": quote.revisions.order_by("revision"),
        "timeline": _commercial_timeline(quote),
    }
    if quote.status == "awarded":
        context["trace"] = traceability(quote)
    return render(request, "web/quotation_detail.html", context)


@login_required
@xframe_options_sameorigin  # so the review page can preview it in an <iframe>
def quotation_pdf(request, pk):
    quote = get_object_or_404(Quotation.objects.all().prefetch_related("lines"), pk=pk)
    pdf = quotation_pdf_bytes(quote)
    # The on-screen preview (inline) is available to anyone who may view the
    # quotation; taking a copy (attachment) needs the download permission.
    inline = bool(request.GET.get("inline"))
    if not inline and not request.user.has_perm_code("quotes.download"):
        return HttpResponseForbidden("You do not have permission to download this.")
    pdf = quotation_pdf_bytes(quote)
    resp = HttpResponse(pdf, content_type="application/pdf")
    disposition = "inline" if inline else "attachment"
    resp["Content-Disposition"] = f'{disposition}; filename="{quote.number}.pdf"'
    if not inline:                       # log real downloads, not preview renders
        from apps.core.audit import audit
        audit(request, "quotation.pdf_downloaded", entity=quote)
    return resp


@login_required
def quotation_excel(request, pk):
    """Export the quotation's items to a real .xlsx — the line table plus totals,
    for a customer or buyer who wants to work with the numbers in a spreadsheet.
    Selling price only: cost and margin never leave the building (Golden Rule)."""
    import io

    import openpyxl
    from openpyxl.styles import Font

    if not request.user.has_perm_code("quotes.download"):
        return HttpResponseForbidden("You do not have permission to export this.")
    quote = get_object_or_404(Quotation.objects.all().prefetch_related("lines"), pk=pk)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    ws.append([f"Quotation {quote.number}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([quote.client_name, "", "", "", quote.title])
    ws.append([])
    header = ["#", "Description", "Qty", "Unit", "Unit price", "Line total"]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for ln in quote.lines.all():
        ws.append([ln.position, ln.description, float(ln.qty), ln.unit,
                   float(ln.unit_price), float(ln.line_total)])
    ws.append([])
    # On an exclusive quotation VAT is added on the invoice, so it reads 0 here.
    vat_on_quote = quote.vat_amount if quote.vat_mode == "inclusive" else 0
    ws.append(["", "", "", "", "Subtotal", float(quote.subtotal)])
    ws.append(["", "", "", "", "VAT", float(vat_on_quote)])
    total_row = ["", "", "", "", "Total", float(quote.total)]
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for col, width in {"A": 6, "B": 48, "C": 8, "D": 10, "E": 14, "F": 14}.items():
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{quote.number}.xlsx"'
    from apps.core.audit import audit
    audit(request, "quotation.excel_exported", entity=quote)
    return resp


@login_required
@require_POST
def quotation_po_extract(request, pk):
    """Read an uploaded purchase order and return its fields as JSON so the PO
    form fills itself in — the estimator confirms rather than retypes. Reuses the
    shared Document Intelligence service; Gemini fills what the patterns miss.
    Stateless — it reads the file, it saves nothing."""
    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("quotes.create"):
        return JsonResponse({}, status=403)
    from apps.knowledge.document_intelligence import (
        extract_po_fields,
        extract_text_from_upload,
    )
    f = request.FILES.get("document")
    text = extract_text_from_upload(f) if f else ""
    fields = extract_po_fields(text, company=quote.company, user=request.user,
                               use_ai=True)
    if not fields.get("value") and quote.total:
        fields["value"] = f"{quote.total:.2f}"     # sensible default: the quote total
    return JsonResponse(fields)


# ── RFQ (the front door: upload → extract → review → approve → quotation) ─────

@login_required
def rfq_list(request):
    return render(request, "web/rfq.html", {
        "rfqs": RFQDocument.objects.all(),
        "can_upload": request.user.has_perm_code("rfq.upload"),
    })


@login_required
@require_POST
def rfq_upload(request):
    if not request.user.has_perm_code("rfq.upload"):
        messages.error(request, "You do not have permission to upload RFQs.")
        return redirect("web:rfq")
    upload = request.FILES.get("file")
    pasted = request.POST.get("text", "").strip()

    if upload is not None:
        rfq = ingest_rfq(request.user.active_company, request.user,
                         uploaded_file=upload, original_name=upload.name)
    elif pasted:
        try:
            rfq = ingest_rfq_text(
                request.user.active_company, request.user, text=pasted,
                original_name=request.POST.get("label", ""))
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("web:rfq")
    else:
        messages.error(request, "Attach a document or paste the RFQ text.")
        return redirect("web:rfq")

    messages.success(request, f"Extracted {rfq.fields.count()} field(s) and "
                              f"{rfq.lines.count()} line(s). Review below.")
    return redirect("web:rfq_detail", pk=rfq.id)


@login_required
def rfq_detail(request, pk):
    rfq = get_object_or_404(
        RFQDocument.objects.all().prefetch_related("fields", "lines"), pk=pk)
    editable = (rfq.status not in ("approved", "rejected")
                and request.user.has_perm_code("rfq.upload"))
    context = {
        "rfq": rfq,
        "can_edit": editable,
        "can_approve": (rfq.status not in ("approved", "rejected")
                        and request.user.has_perm_code("rfq.approve")),
    }
    return render(request, "web/rfq_detail.html", context)


@login_required
@require_POST
def rfq_line(request, pk):
    """Add, edit or remove a line by hand — the reviewer has the last word."""
    rfq = get_object_or_404(RFQDocument.objects.all(), pk=pk)
    if not request.user.has_perm_code("rfq.upload"):
        messages.error(request, "You do not have permission to edit this RFQ.")
        return redirect("web:rfq_detail", pk=pk)
    if rfq.status in ("approved", "rejected"):
        messages.error(request, "This RFQ is closed — it can no longer be edited.")
        return redirect("web:rfq_detail", pk=pk)

    action = request.POST.get("action", "add")
    if action == "add":
        description = request.POST.get("description", "").strip()
        if not description:
            messages.error(request, "Enter a description.")
        else:
            add_line_item(rfq, request.user, description=description,
                          qty=_decimal_or_none(request.POST.get("qty")) or 1,
                          unit=request.POST.get("unit", "each").strip() or "each",
                          unit_price=_decimal_or_none(request.POST.get("unit_price")))
            messages.success(request, "Line added.")
    else:
        line = get_object_or_404(RFQLineItem.objects.filter(rfq=rfq),
                                 pk=request.POST.get("line"))
        if action == "delete":
            delete_line_item(line)
            messages.success(request, "Line removed.")
        else:
            update_line_item(
                line, request.user,
                description=request.POST.get("description"),
                qty=_decimal_or_none(request.POST.get("qty")),
                unit=request.POST.get("unit"),
                unit_price=_decimal_or_none(request.POST.get("unit_price")))
            messages.success(request, "Line updated.")
    return redirect("web:rfq_detail", pk=pk)


@login_required
@require_POST
def rfq_approve(request, pk):
    """Human approval (never automatic) → creates the Quotation + Project DNA."""
    if not request.user.has_perm_code("rfq.approve"):
        messages.error(request, "You do not have permission to approve RFQs.")
        return redirect("web:rfq_detail", pk=pk)
    rfq = get_object_or_404(RFQDocument.objects.all(), pk=pk)
    client_name = request.POST.get("client_name", "").strip()
    if not client_name:
        messages.error(request, "Enter the client name to create the quotation.")
        return redirect("web:rfq_detail", pk=pk)
    approve_rfq(rfq, request.user, client_name=client_name)
    rfq.refresh_from_db()
    messages.success(request, f"Approved → quotation {rfq.quotation.number} created. "
                              "Next: build an estimate, then award to a project.")
    return redirect("web:rfq_detail", pk=pk)


# ── Estimates ─────────────────────────────────────────────────────────────────

@login_required
def estimates_list(request):
    estimates = Estimate.objects.all().order_by("-created_at")
    return render(request, "web/estimates.html",
                  {"estimates": estimates, "can_view_money": _can_view_money(request.user)})


@login_required
def estimate_detail(request, pk):
    estimate = get_object_or_404(Estimate.objects.all().prefetch_related("sections__lines"), pk=pk)
    can_approve = (estimate.status in (EstimateStatus.REVIEW, EstimateStatus.AWAITING_APPROVAL)
                   and request.user.has_perm_code("estimating.approve"))
    return render(request, "web/estimate_detail.html", {
        "estimate": estimate,
        "can_view_money": _can_view_money(request.user),
        "can_approve": can_approve,
        "can_revise": (estimate.status != EstimateStatus.SUPERSEDED
                       and request.user.has_perm_code("estimating.manage")),
    })


@login_required
@require_POST
def estimate_approve(request, pk):
    if not request.user.has_perm_code("estimating.approve"):
        messages.error(request, "You do not have permission to approve estimates.")
        return redirect("web:estimate_detail", pk=pk)
    estimate = get_object_or_404(Estimate.objects.all(), pk=pk)
    approve_estimate(estimate, request.user)
    messages.success(request, f"Estimate {estimate.number} approved.")
    return redirect("web:estimate_detail", pk=pk)


# ── Procurement ───────────────────────────────────────────────────────────────

@login_required
def suppliers_list(request):
    """Suppliers (including any learned from receipts). A search matches supplier
    names AND items we've bought, so 'pipes' answers 'who do we buy pipes from?'"""
    from apps.procurement.models import SupplierPrice

    q = (request.GET.get("q") or "").strip()
    suppliers = Supplier.objects.all()
    if q:
        suppliers = suppliers.filter(name__icontains=q)
    suppliers = suppliers.order_by("-performance_score", "name")

    item_matches = []
    if q:
        # Match each word, and tolerate simple plurals ("pipes" finds "pipe"),
        # so a natural search lands on what we actually bought.
        from django.db.models import Q
        terms = {w for w in q.split() if len(w) > 2}
        terms |= {w[:-1] for w in list(terms) if w.endswith("s")}
        cond = Q()
        for term in terms or {q}:
            cond |= Q(description__icontains=term)
        seen = set()
        for p in (SupplierPrice.objects.filter(cond)
                  .select_related("supplier").order_by("item_key", "-date")):
            key = (p.supplier_id, p.item_key)
            if key not in seen:          # latest price per supplier+item
                seen.add(key)
                item_matches.append(p)
    return render(request, "web/suppliers.html", {
        "suppliers": suppliers, "q": q, "item_matches": item_matches,
        "can_manage": request.user.has_perm_code("procurement.manage")})


@login_required
def supplier_detail(request, pk):
    """One supplier: what we've bought from them (price ledger), the receipts
    that built the record, and the documents kept against them."""
    supplier = get_object_or_404(Supplier.objects.all(), pk=pk)
    prices = supplier.prices.all().order_by("-date")[:200]
    receipts = supplier.receipts.select_related("task").order_by("-reported_at")[:100]
    return render(request, "web/supplier_detail.html", {
        "supplier": supplier, "prices": prices, "receipts": receipts,
        "documents": supplier.documents.all(),
        "can_manage": request.user.has_perm_code("procurement.manage"),
    })


def _supplier_fields(post):
    """Read the supplier form fields; categories arrive comma-separated."""
    cats = [c.strip() for c in (post.get("categories") or "").split(",") if c.strip()]
    return {
        "name": (post.get("name") or "").strip(),
        "contact_person": (post.get("contact_person") or "").strip(),
        "email": (post.get("email") or "").strip(),
        "phone": (post.get("phone") or "").strip(),
        "payment_terms": (post.get("payment_terms") or "credit").strip(),
        "categories": cats,
        "notes": (post.get("notes") or "").strip(),
    }


@login_required
@require_POST
def supplier_create(request):
    """Add a supplier by hand."""
    from apps.core.audit import audit
    if not request.user.has_perm_code("procurement.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:suppliers")
    fields = _supplier_fields(request.POST)
    if not fields["name"]:
        messages.error(request, "A supplier needs a name.")
        return redirect("web:suppliers")
    existing = Supplier.objects.filter(name__iexact=fields["name"]).first()
    if existing:
        messages.info(request, f"{existing.name} is already in your suppliers.")
        return redirect("web:supplier_detail", pk=existing.id)
    supplier = Supplier.objects.create(
        company=request.user.active_company, created_by=request.user,
        updated_by=request.user, **fields)
    audit(request, "supplier.created", entity=supplier)
    messages.success(request, f"Added {supplier.name}.")
    return redirect("web:supplier_detail", pk=supplier.id)


@login_required
@require_POST
def supplier_edit(request, pk):
    """Edit a supplier's details."""
    supplier = get_object_or_404(Supplier.objects.all(), pk=pk)
    if not request.user.has_perm_code("procurement.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:supplier_detail", pk=pk)
    fields = _supplier_fields(request.POST)
    if not fields["name"]:
        messages.error(request, "A supplier needs a name.")
        return redirect("web:supplier_detail", pk=pk)
    for key, value in fields.items():
        setattr(supplier, key, value)
    supplier.preferred = bool(request.POST.get("preferred"))
    supplier.updated_by = request.user
    supplier.save()
    messages.success(request, "Supplier updated.")
    return redirect("web:supplier_detail", pk=pk)


@login_required
@require_POST
def supplier_document(request, pk):
    """Attach a document (quote, certificate, banking, invoice) to a supplier."""
    from apps.core.uploads import validate_upload
    from apps.procurement.models import SupplierDocument

    supplier = get_object_or_404(Supplier.objects.all(), pk=pk)
    if not request.user.has_perm_code("procurement.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:supplier_detail", pk=pk)
    upload = request.FILES.get("file")
    if upload is None:
        messages.error(request, "Choose a file to upload.")
        return redirect("web:supplier_detail", pk=pk)
    try:
        validate_upload(upload)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("web:supplier_detail", pk=pk)
    SupplierDocument.objects.create(
        company=request.user.active_company, supplier=supplier, file=upload,
        name=(request.POST.get("name") or upload.name).strip(),
        doc_type=(request.POST.get("doc_type") or "other").strip(),
        content_type=getattr(upload, "content_type", ""), size_bytes=upload.size,
        created_by=request.user, updated_by=request.user)
    messages.success(request, "Document uploaded.")
    return redirect("web:supplier_detail", pk=pk)


@login_required
@require_POST
def supplier_import(request):
    """Upload an old invoice and extract it for review before it seeds the
    Suppliers database (human confirms first — nothing saved yet)."""
    from apps.core.uploads import validate_upload
    from apps.knowledge.document_intelligence import (
        extract_po_fields,
        extract_text_from_upload,
    )
    if not request.user.has_perm_code("procurement.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:suppliers")
    upload = request.FILES.get("file")
    if upload is None:
        messages.error(request, "Choose an invoice to upload.")
        return redirect("web:suppliers")
    try:
        validate_upload(upload)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("web:suppliers")
    text = extract_text_from_upload(upload)
    fields = extract_po_fields(text, company=request.user.active_company,
                               user=request.user, use_ai=True)
    return render(request, "web/supplier_import_review.html", {
        "supplier_name": fields.get("contact", ""),
        "invoice_number": fields.get("po_number", ""),
        "doc_date": fields.get("po_date", ""),
        "lines": fields.get("lines", []),
        "filename": upload.name,
    })


@login_required
@require_POST
def supplier_import_confirm(request):
    """Save the reviewed old-invoice lines into the Suppliers database."""
    from apps.core.audit import audit
    from apps.procurement.services import learn_from_receipt

    if not request.user.has_perm_code("procurement.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:suppliers")
    name = (request.POST.get("supplier_name") or "").strip()
    if not name:
        messages.error(request, "Enter the supplier's name.")
        return redirect("web:suppliers")
    descriptions = request.POST.getlist("description")
    units = request.POST.getlist("unit")
    prices = request.POST.getlist("unit_price")
    items = [
        {"description": d.strip(), "unit": (units[i] if i < len(units) else "each"),
         "unit_price": (prices[i] if i < len(prices) else "0")}
        for i, d in enumerate(descriptions) if d.strip()
    ]
    doc_date = request.POST.get("doc_date") or None
    supplier, n, created = learn_from_receipt(
        request.user.active_company, request.user,
        supplier_name=name, items=items, date=doc_date or None)
    audit(request, "supplier.imported", entity=supplier)
    verb = "Added" if created else "Updated"
    messages.success(request, f"{verb} {supplier.name} — {n} price(s) recorded.")
    return redirect("web:supplier_detail", pk=supplier.id)


@login_required
def products_list(request):
    """Products we buy, built from the price ledger — with a spend-by-category
    breakdown. Search matches names and aliases."""
    from apps.procurement.models import ProductCategory
    from apps.procurement.services import products_overview, spend_by_category

    q = (request.GET.get("q") or "").strip()
    category = (request.GET.get("category") or "").strip()
    return render(request, "web/products.html", {
        "rows": products_overview(request.user.active_company, q=q, category=category),
        "spend": spend_by_category(request.user.active_company),
        "categories": ProductCategory.choices, "q": q, "category": category,
    })


@login_required
def product_detail(request, pk):
    """The product-knowledge page: who sells it, who's cheapest, how often/when we
    bought it, avg/low/high, price trend, aliases and category."""
    from apps.procurement.models import Product, ProductCategory
    from apps.procurement.services import product_intelligence

    product = get_object_or_404(Product.objects.all(), pk=pk)
    ctx = product_intelligence(product)
    ctx.update({
        "categories": ProductCategory.choices,
        "aliases": product.aliases.all(),
        "other_products": Product.objects.exclude(pk=pk).order_by("name")[:500],
        "can_manage": request.user.has_perm_code("procurement.manage"),
    })
    return render(request, "web/product_detail.html", ctx)


@login_required
@require_POST
def product_edit(request, pk):
    from apps.procurement.models import Product

    product = get_object_or_404(Product.objects.all(), pk=pk)
    if not request.user.has_perm_code("procurement.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:product_detail", pk=pk)
    name = (request.POST.get("name") or "").strip()
    if name:
        product.name = name
    product.category = (request.POST.get("category") or "").strip()
    product.updated_by = request.user
    product.save()
    messages.success(request, "Product updated.")
    return redirect("web:product_detail", pk=pk)


@login_required
@require_POST
def product_alias(request, pk):
    from apps.procurement.models import Product
    from apps.procurement.services import add_product_alias

    product = get_object_or_404(Product.objects.all(), pk=pk)
    if not request.user.has_perm_code("procurement.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:product_detail", pk=pk)
    label = (request.POST.get("label") or "").strip()
    if label:
        add_product_alias(product, request.user, label)
        messages.success(request, f"“{label}” now maps to {product.name}.")
    return redirect("web:product_detail", pk=pk)


@login_required
@require_POST
def product_merge(request, pk):
    """Fold another product into this one (duplicate cleanup, #8)."""
    from apps.procurement.models import Product
    from apps.procurement.services import merge_products

    keep = get_object_or_404(Product.objects.all(), pk=pk)
    if not request.user.has_perm_code("procurement.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:product_detail", pk=pk)
    drop = Product.objects.filter(pk=request.POST.get("drop")).first()
    if drop and drop.id != keep.id:
        name = drop.name
        merge_products(keep, drop, request.user)
        messages.success(request, f"Merged “{name}” into {keep.name}.")
    return redirect("web:product_detail", pk=pk)


@login_required
def requests_list(request):
    """Internal procurement requests — what tasks need, before it's bought."""
    from apps.procurement.models import ProcurementRequest, ProcurementRequestStatus
    from apps.procurement.services import procurement_approval_required

    status = (request.GET.get("status") or "").strip()
    qs = ProcurementRequest.objects.all().select_related("task", "project", "requested_by")
    if status:
        qs = qs.filter(status=status)
    awaiting = ProcurementRequest.objects.filter(
        status=ProcurementRequestStatus.SUBMITTED).count()
    return render(request, "web/requests.html", {
        "requests": qs, "statuses": ProcurementRequestStatus.choices, "status": status,
        "awaiting": awaiting,
        "approval_required": procurement_approval_required(request.user.active_company),
        "can_approve": request.user.has_perm_code("procurement.manage"),
        "can_config": request.user.has_perm_code("company.manage"),
    })


@login_required
def request_new(request):
    """Create a procurement request. GET renders the form; POST creates it and
    submits it (auto-approved unless the company requires approval)."""
    from apps.execution.models import Task
    from apps.procurement.services import create_request, submit_request

    if request.method == "POST":
        task = Task.objects.filter(pk=request.POST.get("task")).first()
        descriptions = request.POST.getlist("description")
        quantities = request.POST.getlist("quantity")
        units = request.POST.getlist("unit")
        prices = request.POST.getlist("est_unit_price")
        lines = [
            {"description": d.strip(),
             "quantity": quantities[i] if i < len(quantities) else 1,
             "unit": units[i] if i < len(units) else "each",
             "est_unit_price": prices[i] if i < len(prices) else ""}
            for i, d in enumerate(descriptions) if d.strip()
        ]
        if not lines:
            messages.error(request, "Add at least one item.")
            return redirect("web:request_new")
        req = create_request(
            request.user.active_company, request.user,
            title=(request.POST.get("title") or "Materials request").strip(),
            task=task, notes=(request.POST.get("notes") or "").strip(),
            needed_by=request.POST.get("needed_by") or None, lines=lines)
        submit_request(req, request.user)
        messages.success(request, f"Request {req.number} created.")
        return redirect("web:request_detail", pk=req.id)

    task = Task.objects.filter(pk=request.GET.get("task")).first()
    return render(request, "web/request_new.html", {
        "task": task, "tasks": Task.objects.all().order_by("name")[:500]})


@login_required
def request_detail(request, pk):
    from apps.procurement.models import ProcurementRequest

    req = get_object_or_404(
        ProcurementRequest.objects.select_related("task", "project", "requested_by",
                                                  "approved_by"), pk=pk)
    return render(request, "web/request_detail.html", {
        "req": req, "lines": req.lines.all(),
        "can_approve": request.user.has_perm_code("procurement.manage"),
    })


def _request_guard(request, pk, *, need_manage=False):
    from apps.procurement.models import ProcurementRequest
    req = get_object_or_404(ProcurementRequest.objects.all(), pk=pk)
    if need_manage and not request.user.has_perm_code("procurement.manage"):
        messages.error(request, "You do not have permission.")
        return req, False
    return req, True


@login_required
@require_POST
def request_approve(request, pk):
    from apps.procurement.services import approve_request
    req, ok = _request_guard(request, pk, need_manage=True)
    if ok:
        approve_request(req, request.user)
        messages.success(request, f"Approved {req.number}.")
    return redirect("web:request_detail", pk=pk)


@login_required
@require_POST
def request_reject(request, pk):
    from apps.procurement.services import reject_request
    req, ok = _request_guard(request, pk, need_manage=True)
    if ok:
        reject_request(req, request.user, reason=request.POST.get("reason", ""))
        messages.info(request, f"Rejected {req.number}.")
    return redirect("web:request_detail", pk=pk)


@login_required
@require_POST
def request_fulfil(request, pk):
    from apps.procurement.services import fulfil_request
    req, ok = _request_guard(request, pk, need_manage=True)
    if ok:
        fulfil_request(req, request.user)
        messages.success(request, f"{req.number} marked as purchased.")
    return redirect("web:request_detail", pk=pk)


@login_required
@require_POST
def request_settings(request):
    """Toggle whether purchases need approval (optional per company)."""
    from apps.administration.models import CompanySettings
    if not request.user.has_perm_code("company.manage"):
        messages.error(request, "You do not have permission.")
        return redirect("web:requests")
    s, _ = CompanySettings.objects.get_or_create(company=request.user.active_company)
    rules = dict(s.approval_rules or {})
    rules["procurement_required"] = bool(request.POST.get("procurement_required"))
    s.approval_rules = rules
    s.save(update_fields=["approval_rules", "updated_at"])
    messages.success(request, "Procurement approval setting updated.")
    return redirect("web:requests")


@login_required
def procurement_dashboard(request):
    """The Procurement home page — suppliers, purchases, requests and spend at a
    glance, with the section's sub-navigation."""
    from apps.procurement.services import procurement_dashboard_metrics
    return render(request, "web/procurement_dashboard.html",
                  procurement_dashboard_metrics(request.user.active_company))


@login_required
def procurement_prices(request):
    """Price history — the append-only ledger of everything we've paid."""
    from apps.procurement.models import SupplierPrice
    q = (request.GET.get("q") or "").strip()
    rows = SupplierPrice.objects.select_related("supplier", "product")
    if q:
        rows = rows.filter(description__icontains=q)
    rows = rows.order_by("-date", "-created_at")[:300]
    return render(request, "web/procurement_prices.html", {"rows": rows, "q": q})


@login_required
def purchase_orders_list(request):
    pos = PurchaseOrder.objects.all().select_related("supplier").prefetch_related("lines")
    return render(request, "web/purchase_orders.html",
                  {"purchase_orders": pos, "can_view_money": _can_view_money(request.user)})


@login_required
def po_detail(request, pk):
    po = get_object_or_404(
        PurchaseOrder.objects.all().select_related("supplier").prefetch_related("lines"), pk=pk)
    can_approve = (po.status in ("draft", "pending_approval")
                   and request.user.has_perm_code("po.approve"))
    return render(request, "web/po_detail.html", {
        "po": po,
        "match": three_way_match(po),
        "can_view_money": _can_view_money(request.user),
        "can_approve": can_approve,
        "can_receive": request.user.has_perm_code("procurement.manage"),
    })


@login_required
@require_POST
def po_approve(request, pk):
    if not request.user.has_perm_code("po.approve"):
        messages.error(request, "You do not have permission to approve purchase orders.")
        return redirect("web:po_detail", pk=pk)
    po = get_object_or_404(PurchaseOrder.objects.all(), pk=pk)
    po.status = "approved"
    po.approved_by = request.user
    po.save(update_fields=["status", "approved_by"])
    messages.success(request, f"Purchase order {po.number} approved.")
    return redirect("web:po_detail", pk=pk)


# ── Commercial (finance only) ─────────────────────────────────────────────────

@login_required
def commercial(request):
    if not _can_view_money(request.user):
        messages.error(request, "Commercial data requires the finance permission.")
        return redirect("web:dashboard")
    dash = commercial_dashboard(request.user.active_company)
    aging = dash["aging"]
    # "90+" can't be reached via template dot-notation, so pre-build ordered rows.
    aging_rows = [("Current", aging["current"]), ("30 days", aging["30"]),
                  ("60 days", aging["60"]), ("90+ days", aging["90+"])]
    return render(request, "web/commercial.html", {
        "commercial": dash,
        "aging_rows": aging_rows,
        "invoices": Invoice.objects.all().select_related("project").prefetch_related(
            "lines", "payments"),
        "can_finance": request.user.has_perm_code("finance.manage"),
    })


# ── Lulama (AI orchestrator) ──────────────────────────────────────────────────

@login_required
def lulama(request):
    if not request.user.has_perm_code("ai.generate"):
        messages.error(request, "AI features require the ai.generate permission.")
        return redirect("web:dashboard")
    context = {"projects": Project.objects.all(), "request_text": "Prepare this project"}
    if request.method == "POST":
        project = None
        pid = request.POST.get("project")
        if pid:
            project = get_object_or_404(Project.objects.all(), pk=pid)
        text = request.POST.get("request", "").strip() or "Give me a status overview"
        context["request_text"] = text
        context["selected_project"] = pid
        context["interaction"] = orchestrate(
            request.user.active_company, request.user, text, project=project)
    return render(request, "web/lulama.html", context)


# ── Operating actions (managers actually do the work here) ────────────────────

def _to_decimal(raw, default="0"):
    try:
        return Decimal(str(raw or default))
    except (InvalidOperation, TypeError):
        return Decimal(default)


@login_required
@require_POST
def compliance_item_approve(request, pk):
    """Approve a compliance item from the project page — the gate recomputes live."""
    if not request.user.has_perm_code("compliance.override"):
        messages.error(request, "You do not have permission to approve compliance items.")
        return redirect("web:dashboard")
    item = get_object_or_404(ComplianceItem.objects.all().select_related("project"), pk=pk)
    expiry = request.POST.get("expiry") or None
    approve_item(item, request.user, expiry=expiry)
    messages.success(request, f"Approved: {item.name}.")
    return redirect("web:project_detail", pk=item.project_id)


@login_required
@require_POST
def project_override(request, pk):
    """Authorised override of the compliance gate — reason required, audited."""
    if not request.user.has_perm_code("compliance.override"):
        messages.error(request, "You do not have permission to override compliance.")
        return redirect("web:project_detail", pk=pk)
    project = get_object_or_404(Project.objects.all(), pk=pk)
    reason = request.POST.get("reason", "").strip()
    if not reason:
        messages.error(request, "An override needs a reason.")
        return redirect("web:project_detail", pk=pk)
    override_gate(project, request.user, reason=reason)
    messages.success(request, "Compliance gate overridden (audited).")
    return redirect("web:project_detail", pk=pk)


@login_required
@require_POST
def po_receive(request, pk):
    """Record goods received (a GRN) against a PO — feeds the 3-way match."""
    if not request.user.has_perm_code("procurement.manage"):
        messages.error(request, "You do not have permission to receive goods.")
        return redirect("web:po_detail", pk=pk)
    po = get_object_or_404(PurchaseOrder.objects.all().prefetch_related("lines"), pk=pk)
    company = request.user.active_company
    grn = GRN.objects.create(company=company, purchase_order=po, seq=po.grns.count() + 1,
                             date=timezone.localdate(), received_by=request.user,
                             created_by=request.user, updated_by=request.user)
    received = 0
    for line in po.lines.all():
        qty = _to_decimal(request.POST.get(f"qty_{line.id}"))
        if qty > 0:
            GRNLine.objects.create(company=company, grn=grn, po_line=line,
                                   description=line.description, qty_received=qty,
                                   created_by=request.user, updated_by=request.user)
            received += 1
    if received:
        messages.success(request, f"Goods received on {po.number} ({received} line(s)).")
    else:
        grn.delete(hard=True)
        messages.error(request, "Enter a received quantity on at least one line.")
    return redirect("web:po_detail", pk=pk)


@login_required
@require_POST
def estimate_revise(request, pk):
    """Create a new estimate revision (the prior one is superseded, never overwritten)."""
    if not request.user.has_perm_code("estimating.manage"):
        messages.error(request, "You do not have permission to revise estimates.")
        return redirect("web:estimate_detail", pk=pk)
    estimate = get_object_or_404(Estimate.objects.all(), pk=pk)
    new = create_revision(estimate, request.user, reason=request.POST.get("reason", "").strip())
    messages.success(request, f"Created revision v{new.version}.")
    return redirect("web:estimate_detail", pk=new.id)


@login_required
@require_POST
def project_progress_claim(request, pk):
    """Raise a progress claim (a % of the contract value, with retention held)."""
    if not request.user.has_perm_code("finance.manage"):
        messages.error(request, "You do not have permission to raise claims.")
        return redirect("web:project_detail", pk=pk)
    project = get_object_or_404(Project.objects.all(), pk=pk)
    pct = _to_decimal(request.POST.get("percent_complete"))
    retention = _to_decimal(request.POST.get("retention"), "10")
    invoice = create_progress_claim(project, request.user, percent_complete=pct,
                                    retention_pct=retention)
    messages.success(request, f"Progress claim {invoice.number} raised (draft).")
    return redirect("web:project_detail", pk=pk)


@login_required
@require_POST
def invoice_payment(request, pk):
    """Record a customer payment against an invoice (POP)."""
    if not request.user.has_perm_code("finance.manage"):
        messages.error(request, "You do not have permission to record payments.")
        return redirect("web:commercial")
    invoice = get_object_or_404(Invoice.objects.all(), pk=pk)
    amount = _to_decimal(request.POST.get("amount"))
    if amount <= 0:
        messages.error(request, "Enter a payment amount.")
        return redirect("web:commercial")
    record_payment(invoice, request.user, amount=amount,
                   reference=request.POST.get("reference", ""))
    messages.success(request, f"Payment of R{amount} recorded on {invoice.number}.")
    return redirect("web:commercial")


# ══════════════════════════════════════════════════════════════════════════════
# Customers — client ORGANISATIONS, not names in a text field
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def customers_list(request):
    from apps.customers.models import Customer
    from apps.customers.services import customer_overview

    query = request.GET.get("q", "").strip()
    customers = Customer.objects.all().prefetch_related("contacts", "departments")
    if query:
        customers = customers.filter(name__icontains=query)
    status = request.GET.get("status")
    if status:
        customers = customers.filter(status=status)

    rows = [{"customer": c, "stats": customer_overview(c)} for c in customers]
    return render(request, "web/customers.html", {
        "rows": rows, "q": query, "status": status,
        "statuses": Customer._meta.get_field("status").choices,
        "can_manage": request.user.has_perm_code("projects.create"),
    })


@login_required
def customer_detail(request, pk):
    from apps.customers.models import Customer, CustomerContact, RESPONSIBILITIES
    from apps.customers.services import (
        customer_overview,
        responsibility_matrix,
        route_document,
    )
    from apps.projects.models import Project
    from apps.quotes.models import Quotation

    customer = get_object_or_404(
        Customer.objects.prefetch_related("departments", "branches", "sites",
                                          "contacts", "contracts"), pk=pk)

    # Contacts grouped by department — an org chart, not a flat list.
    grouped, unassigned = [], []
    for dept in customer.departments.all():
        people = [c for c in customer.contacts.all() if c.department_id == dept.id]
        if people:
            grouped.append({"department": dept, "contacts": people})
    unassigned = [c for c in customer.contacts.all() if c.department_id is None]

    return render(request, "web/customer_detail.html", {
        "customer": customer,
        "stats": customer_overview(customer),
        "grouped": grouped,
        "unassigned": unassigned,
        "departments": customer.departments.all(),
        "matrix": responsibility_matrix(customer),
        "routing": [route_document(customer, k) for k in
                    ("quotation", "invoice", "progress_report", "safety_file")],
        "roles": __import__("apps.customers.models", fromlist=["CONTACT_ROLES"]).CONTACT_ROLES,
        "responsibilities": RESPONSIBILITIES.items(),
        "methods": CustomerContact.Method.choices,
        "quotations": Quotation.objects.filter(customer=customer)[:10],
        "projects": Project.objects.filter(customer=customer)[:10],
        "can_manage": request.user.has_perm_code("projects.create"),
    })


#: The customer fields the edit form may set — a whitelist, so a stray POST key
#: can never write to something it shouldn't.
_CUSTOMER_EDIT_FIELDS = (
    "name", "trading_name", "registration_no", "vat_no", "tax_no", "industry",
    "province", "city", "physical_address", "postal_address", "postal_code",
    "telephone", "mobile", "email", "vendor_number", "vendor_portal",
    "vendor_note", "payment_terms_note", "status",
)


@login_required
def customer_edit(request, pk):
    """Edit a customer's own details — identity, addresses, VAT/registration
    numbers, the vendor number we trade under with them, and payment terms."""
    from apps.customers.models import Customer, CustomerStatus

    customer = get_object_or_404(Customer.objects.all(), pk=pk)
    if not request.user.has_perm_code("projects.create"):
        messages.error(request, "You do not have permission to edit customers.")
        return redirect("web:customer_detail", pk=pk)

    if request.method == "POST":
        if not request.POST.get("name", "").strip():
            messages.error(request, "A company name is required.")
        else:
            for field in _CUSTOMER_EDIT_FIELDS:
                if field in request.POST:
                    setattr(customer, field, request.POST.get(field, "").strip())
            terms = request.POST.get("payment_terms_days", "").strip()
            if terms.isdigit():
                customer.payment_terms_days = int(terms)
            customer.updated_by = request.user
            customer.save()
            messages.success(request, f"{customer.display_name} updated.")
            return redirect("web:customer_detail", pk=pk)

    return render(request, "web/customer_edit.html", {
        "customer": customer,
        "statuses": CustomerStatus.choices,
    })


@login_required
@require_POST
def customer_create(request):
    from apps.customers.services import create_customer
    if not request.user.has_perm_code("projects.create"):
        messages.error(request, "You do not have permission to add customers.")
        return redirect("web:customers")
    name = request.POST.get("name", "").strip()
    if not name:
        messages.error(request, "A company name is required.")
        return redirect("web:customers")
    customer = create_customer(
        request.user.active_company, request.user, name=name,
        industry=request.POST.get("industry", "").strip(),
        email=request.POST.get("email", "").strip(),
        telephone=request.POST.get("telephone", "").strip(),
        city=request.POST.get("city", "").strip(),
        vendor_number=request.POST.get("vendor_number", "").strip(),
    )
    messages.success(request, f"{customer.name} added ({customer.code}) with a "
                              "standard department structure.")
    return redirect("web:customer_detail", pk=customer.id)


@login_required
@require_POST
def customer_contact_save(request, pk):
    """Add or update a person at the customer, with their roles and — the part
    that matters — what they are empowered to do."""
    from apps.customers.models import Customer, CustomerContact, CustomerDepartment
    customer = get_object_or_404(Customer.objects.all(), pk=pk)
    if not request.user.has_perm_code("projects.create"):
        messages.error(request, "You do not have permission.")
        return redirect("web:customer_detail", pk=pk)

    action = request.POST.get("action", "add")
    if action == "delete":
        contact = get_object_or_404(
            CustomerContact.objects.filter(customer=customer),
            pk=request.POST.get("contact"))
        contact.delete()
        messages.success(request, "Contact removed.")
        return redirect("web:customer_detail", pk=pk)

    department = None
    if request.POST.get("department"):
        department = CustomerDepartment.objects.filter(
            customer=customer, pk=request.POST["department"]).first()

    fields = {
        "full_name": request.POST.get("full_name", "").strip(),
        "job_title": request.POST.get("job_title", "").strip(),
        "email": request.POST.get("email", "").strip(),
        "telephone": request.POST.get("telephone", "").strip(),
        "mobile": request.POST.get("mobile", "").strip(),
        "extension": request.POST.get("extension", "").strip(),
        "whatsapp": request.POST.get("whatsapp", "").strip(),
        "preferred_method": request.POST.get("preferred_method", "email"),
        "department": department,
        "roles": request.POST.getlist("roles"),
        "responsibilities": request.POST.getlist("responsibilities"),
    }
    if action == "update":
        contact = get_object_or_404(
            CustomerContact.objects.filter(customer=customer),
            pk=request.POST.get("contact"))
        for key, value in fields.items():
            setattr(contact, key, value)
        contact.status = request.POST.get("status", contact.status)
        contact.is_primary = bool(request.POST.get("is_primary"))
        contact.save()
        messages.success(request, f"{contact.full_name} updated.")
    else:
        if not fields["full_name"]:
            messages.error(request, "A name is required.")
            return redirect("web:customer_detail", pk=pk)
        CustomerContact.objects.create(
            company=request.user.active_company, customer=customer,
            is_primary=bool(request.POST.get("is_primary")),
            created_by=request.user, updated_by=request.user, **fields)
        messages.success(request, f"{fields['full_name']} added.")
    return redirect("web:customer_detail", pk=pk)


@login_required
@require_POST
def customer_department(request, pk):
    from apps.customers.models import Customer, CustomerDepartment
    customer = get_object_or_404(Customer.objects.all(), pk=pk)
    if not request.user.has_perm_code("projects.create"):
        messages.error(request, "You do not have permission.")
    elif request.POST.get("action") == "delete":
        dept = get_object_or_404(CustomerDepartment.objects.filter(customer=customer),
                                 pk=request.POST.get("department"))
        dept.delete()
        messages.success(request, "Department removed.")
    elif request.POST.get("name", "").strip():
        CustomerDepartment.objects.create(
            company=request.user.active_company, customer=customer,
            name=request.POST["name"].strip(),
            email=request.POST.get("email", "").strip(),
            created_by=request.user, updated_by=request.user)
        messages.success(request, "Department added.")
    return redirect("web:customer_detail", pk=pk)


@login_required
def customer_contact_detail(request, pk):
    from apps.customers.models import CustomerContact
    from apps.customers.services import contact_timeline

    contact = get_object_or_404(
        CustomerContact.objects.select_related("customer", "department"), pk=pk)
    return render(request, "web/customer_contact.html", {
        "contact": contact,
        "customer": contact.customer,
        "timeline": contact_timeline(contact),
    })


# ══════════════════════════════════════════════════════════════════════════════
# Quotation builder — the commercial gateway's screens
# ══════════════════════════════════════════════════════════════════════════════

def _quote_guard(request, pk):
    """Fetch the quotation and refuse edits it should not accept.

    Locking is enforced in the service too; this is the polite version that
    redirects with a message instead of raising.
    """
    from apps.quotes.services import QuotationError, guard_editable
    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("quotes.create"):
        messages.error(request, "You do not have permission to edit quotations.")
        return quote, False
    try:
        guard_editable(quote)
    except QuotationError as exc:
        messages.error(request, str(exc))
        return quote, False
    return quote, True


@login_required
def quotation_new(request, pk=None):
    """One guided page, top to bottom, in the order an estimator actually works:
    customer → contact → job type → the basics → scope → items → attachments →
    create. Everything past the basics is optional, so a quick quotation is a
    customer, a type and a title; the depth is there when the job needs it.

    The SAME page edits an existing quotation (pk given): it is prefilled and
    saving updates the quotation in place, so editing is identical to creating.

    RFQ-sourced quotations are created by approving an RFQ (which links the
    source document), so that route lives on the RFQ screen, not here. Copying
    an existing quotation is the one alternative offered, tucked below the fold.
    """
    from apps.customers.models import Customer
    from apps.quotes.models import QuotationType, VatMode
    from apps.quotes.services import (
        QuotationError,
        add_lines_bulk,
        apply_type_template,
        create_quotation,
        duplicate,
        ensure_quotation_types,
        guard_editable,
        parse_grid_lines,
    )

    if not request.user.has_perm_code("quotes.create"):
        messages.error(request, "You do not have permission to create quotations.")
        return redirect("web:quotations")

    company = request.user.active_company
    ensure_quotation_types(company)

    editing = pk is not None
    quote = None
    if editing:
        quote = get_object_or_404(Quotation.objects.all(), pk=pk)
        # A finalized (approved+) quotation is the commercial record — changing it
        # is a revision, not an edit, so the editor is closed.
        if quote.is_finalized:
            messages.error(request, f"{quote.display_number} is read-only — "
                                    "create a revision to change it.")
            return redirect("web:quotation_detail", pk=pk)

    if request.method == "POST":
        method = request.POST.get("method", "blank")

        if method == "copy" and not editing:
            source = get_object_or_404(Quotation.objects.all(),
                                       pk=request.POST.get("source"))
            customer = Customer.objects.filter(
                pk=request.POST.get("customer")).first()
            new = duplicate(source, request.user, customer=customer)
            messages.success(request, f"Copied {source.display_number} → "
                                      f"{new.display_number}.")
            return redirect("web:quotation_detail", pk=new.id)

        # Step 1 is not optional: a quotation is always for a customer.
        customer = Customer.objects.filter(pk=request.POST.get("customer")).first()
        if customer is None:
            messages.error(request, "Choose the customer this quotation is for.")
            return redirect("web:quotation_edit", pk=pk) if editing \
                else redirect("web:quotation_new")

        # ── Edit: update the existing quotation in place ────────────────────
        if editing:
            try:
                guard_editable(quote)
            except QuotationError as exc:
                messages.error(request, str(exc))
                return redirect("web:quotation_detail", pk=quote.id)
            quote.client_name = customer.name
            quote.customer = customer
            quote.title = request.POST.get("title", "").strip()
            quote.site = request.POST.get("site", "").strip()
            quote.quotation_type = QuotationType.objects.filter(
                pk=request.POST.get("quotation_type")).first()
            quote.vat_mode = request.POST.get("vat_mode", VatMode.EXCLUSIVE)
            quote.discount_amount = _positive_decimal(
                request.POST.get("discount_amount"))
            quote.scope_of_work = request.POST.get("scope_of_work", "").strip()
            quote.vendor_number = customer.vendor_number
            _apply_customer_hierarchy(quote, request)
            quote.save()
            _save_quotation_uploads(request, quote, company)
            # The grid is the whole line set — replace it wholesale, same as the
            # create page builds it.
            pasted = request.POST.get("pasted_items", "").strip()
            quote.lines.all().delete()
            if pasted:
                rows = parse_grid_lines(pasted)
                if rows:
                    add_lines_bulk(quote, request.user, rows)
            messages.success(request, f"Quotation {quote.number} updated.")
            return redirect("web:quotation_detail", pk=quote.id)

        # The number is always system-allocated — never an editable field, so it
        # cannot collide or be skipped.
        quote = create_quotation(
            company, request.user,
            client_name=customer.name,
            title=request.POST.get("title", "").strip(),
            site=request.POST.get("site", "").strip(),
        )
        quote.customer = customer
        quote.quotation_type = QuotationType.objects.filter(
            pk=request.POST.get("quotation_type")).first()
        quote.vat_mode = request.POST.get("vat_mode", VatMode.EXCLUSIVE)
        quote.discount_amount = _positive_decimal(
            request.POST.get("discount_amount"))
        quote.scope_of_work = request.POST.get("scope_of_work", "").strip()
        quote.customer_reference = request.POST.get("customer_reference", "").strip()
        # Snapshot the vendor code we trade under with THIS customer.
        quote.vendor_number = customer.vendor_number
        quote.prepared_by = request.user
        _apply_customer_hierarchy(quote, request)
        quote.save()

        # The type decides the shape of the quotation, so its sections are
        # seeded now rather than left for the estimator to remember.
        seeded = apply_type_template(quote, request.user)

        # One upload control handles the scope document and any drawings, BOQ or
        # photos — all attached to the quotation, each kept for next year.
        _save_quotation_uploads(request, quote, company)

        # Items come from the spreadsheet grid, serialised as tab-separated rows.
        # The grid's fourth column is a SELLING price, so use the price-based
        # parser (not the cost+markup one) — otherwise the price lands in
        # unit_cost and the line prints a zero unit price.
        pasted = request.POST.get("pasted_items", "").strip()
        added = 0
        if pasted:
            rows = parse_grid_lines(pasted)
            if rows:
                added = add_lines_bulk(quote, request.user, rows)

        parts = [f"Quotation {quote.number} created"]
        if seeded:
            parts.append(f"{seeded} section(s) for a "
                         f"{quote.quotation_type.label.lower()} job")
        if added:
            parts.append(f"{added} item(s) added")
        messages.success(request, ". ".join(parts) + ".")

        # Remind the estimator to set up standard terms — the quotation prints
        # without a Terms & Conditions section until they do. The nudge stops on
        # its own once terms are saved in the Company Profile.
        from apps.identity.profile import document_terms
        if not document_terms(company, kind="quotation").strip():
            messages.warning(
                request,
                "No quotation terms & conditions are set, so this quotation "
                "prints without them. Add your standard terms under Company "
                "Profile → Commercial document settings and every quotation will "
                "carry them automatically.")

        return redirect("web:quotation_detail", pk=quote.id)

    # Contacts depend on the chosen customer, so the page filters them client
    # side from this map rather than making a round trip on every selection.
    contacts_by_customer = {}
    vendor_by_customer = {}
    for c in Customer.objects.all():
        vendor_by_customer[str(c.id)] = c.vendor_number
        contacts_by_customer[str(c.id)] = [
            {"id": str(ct.id), "name": ct.full_name,
             "role": ct.job_title or (ct.department.name if ct.department_id else ""),
             "email": ct.email, "phone": ct.telephone or ct.mobile}
            for ct in c.contacts.filter(status="active")
        ]

    # When editing, seed the grid with the existing lines (same four fields the
    # grid stores), so the estimator edits exactly what is on the quotation.
    seed_lines = []
    if editing:
        for ln in quote.lines.all():
            seed_lines.append({"description": ln.description, "qty": f"{ln.qty:g}",
                               "unit": ln.unit,
                               "unit_price": f"{ln.effective_unit_price:.2f}"})

    return render(request, "web/quotation_new.html", {
        "customers": Customer.objects.all(),
        "types": QuotationType.objects.all(),
        "vat_modes": VatMode.choices,
        "vat_rate": 15,
        "recent": Quotation.objects.all()[:20],
        # Passed as objects, not strings — json_script serialises them safely.
        "contacts_json": contacts_by_customer,
        "vendor_json": vendor_by_customer,
        "editing": editing,
        "quote": quote,
        "seed_lines": seed_lines,
    })


@login_required
@require_POST
def quotation_extract(request):
    """Live extraction for the create page. Given scope text and/or uploaded
    documents, return candidate line items and related-item suggestions as JSON.

    The page calls this as the estimator types and when a file is dropped, so it
    is deliberately stateless — it reads, it never writes. Nothing here lands on
    a quotation; the estimator accepts what they want in the grid. It reuses the
    one shared Document Intelligence service, the same engine the RFQ front door
    reads with.
    """
    if not request.user.has_perm_code("quotes.create"):
        return JsonResponse({"items": [], "suggestions": []}, status=403)

    from apps.knowledge.document_intelligence import (
        extract_items,
        extract_text_from_upload,
        suggest_related_items,
    )

    from apps.core.uploads import clean_uploads
    type_key = request.POST.get("type_key", "").strip() or None
    files, _ = clean_uploads(request.FILES.getlist("documents"))  # only valid types
    parts = [request.POST.get("scope", "")]
    for f in files:
        parts.append(extract_text_from_upload(f))
    text = "\n".join(p for p in parts if p)

    # A document upload is a deliberate action, so it is worth a metered AI pass
    # (Gemini) for the lines the pattern parser misses. Live typing stays
    # deterministic — instant and free — so it never bills a keystroke.
    use_ai = bool(files)

    return JsonResponse({
        "items": extract_items(text, type_key=type_key,
                               company=request.user.active_company,
                               user=request.user, use_ai=use_ai),
        "suggestions": suggest_related_items(text, request.POST.getlist("existing")),
    })


def _apply_customer_hierarchy(quote, request):
    """Customer → branch → site → department → contact, as far as it was given."""
    from apps.customers.models import (
        CustomerBranch,
        CustomerContact,
        CustomerDepartment,
        CustomerSite,
    )
    if not quote.customer_id:
        return
    quote.branch = CustomerBranch.objects.filter(
        customer=quote.customer, pk=request.POST.get("branch")).first()
    quote.customer_site = CustomerSite.objects.filter(
        customer=quote.customer, pk=request.POST.get("customer_site")).first()
    quote.department = CustomerDepartment.objects.filter(
        customer=quote.customer, pk=request.POST.get("department")).first()
    quote.contact = CustomerContact.objects.filter(
        customer=quote.customer, pk=request.POST.get("contact")).first()


@login_required
@require_POST
def quotation_header(request, pk):
    """Save the header — terms, references, scope, and who it goes to."""
    quote, allowed = _quote_guard(request, pk)
    if not allowed:
        return redirect(_edit_url(pk))

    from apps.customers.models import Customer
    from apps.quotes.models import QuotationType

    for field in ("title", "site", "customer_reference", "rfq_reference",
                  "project_reference", "vendor_number", "scope_of_work",
                  "exclusions", "assumptions", "notes"):
        if field in request.POST:
            setattr(quote, field, request.POST.get(field, "").strip())
    if request.POST.get("customer"):
        customer = Customer.objects.filter(pk=request.POST["customer"]).first()
        if customer:
            quote.customer = customer
            quote.client_name = customer.name
    _apply_customer_hierarchy(quote, request)
    if request.POST.get("vat_mode"):
        quote.vat_mode = request.POST["vat_mode"]
    if request.POST.get("vat_rate"):
        quote.vat_rate = _decimal_or_none(request.POST["vat_rate"]) or quote.vat_rate
    if request.POST.get("quotation_type"):
        quote.quotation_type = QuotationType.objects.filter(
            pk=request.POST["quotation_type"]).first()
    quote.validity_date = request.POST.get("validity_date") or None
    quote.updated_by = request.user
    quote.save()
    messages.success(request, "Quotation updated.")
    return redirect(_edit_url(pk))


@login_required
@require_POST
def quotation_line(request, pk):
    """Add, edit or remove a priced line."""
    from apps.quotes.models import QuotationLine, QuotationSection

    quote, allowed = _quote_guard(request, pk)
    if not allowed:
        return redirect(_edit_url(pk))

    action = request.POST.get("action", "add")
    if action == "delete":
        line = get_object_or_404(QuotationLine.objects.filter(quotation=quote),
                                 pk=request.POST.get("line"))
        line.delete()
        messages.success(request, "Line removed.")
        return redirect(_edit_url(pk))

    section = QuotationSection.objects.filter(
        quotation=quote, pk=request.POST.get("section")).first()
    values = {
        "description": request.POST.get("description", "").strip(),
        "category": request.POST.get("category", "other"),
        "qty": _decimal_or_none(request.POST.get("qty")) or Decimal("1"),
        "unit": request.POST.get("unit", "each").strip() or "each",
        "unit_cost": _decimal_or_none(request.POST.get("unit_cost")) or Decimal("0"),
        "markup_pct": _decimal_or_none(request.POST.get("markup_pct")) or Decimal("0"),
        "discount_pct": _decimal_or_none(request.POST.get("discount_pct")) or Decimal("0"),
        "unit_price": _decimal_or_none(request.POST.get("unit_price")) or Decimal("0"),
        "notes": request.POST.get("notes", "").strip(),
    }

    if action == "update":
        line = get_object_or_404(QuotationLine.objects.filter(quotation=quote),
                                 pk=request.POST.get("line"))
        for key, value in values.items():
            setattr(line, key, value)
        line.section = section
        line.updated_by = request.user
        line.save()
        messages.success(request, "Line updated.")
    else:
        if not values["description"]:
            messages.error(request, "A description is required.")
        else:
            QuotationLine.objects.create(
                company=quote.company, quotation=quote, section=section,
                position=quote.lines.count() + 1,
                created_by=request.user, updated_by=request.user, **values)
            messages.success(request, "Line added.")
    return redirect(_edit_url(pk))


@login_required
@require_POST
def quotation_section(request, pk):
    from apps.quotes.models import QuotationSection
    quote, allowed = _quote_guard(request, pk)
    if not allowed:
        return redirect(_edit_url(pk))
    if request.POST.get("action") == "delete":
        section = get_object_or_404(QuotationSection.objects.filter(quotation=quote),
                                    pk=request.POST.get("section"))
        section.delete()          # lines survive, they just lose their grouping
        messages.success(request, "Section removed — its lines were kept.")
    elif request.POST.get("name", "").strip():
        QuotationSection.objects.create(
            company=quote.company, quotation=quote,
            name=request.POST["name"].strip(),
            position=quote.sections.count() + 1,
            created_by=request.user, updated_by=request.user)
        messages.success(request, "Section added.")
    return redirect(_edit_url(pk))


@login_required
@require_POST
def quotation_transition(request, pk):
    from apps.quotes.services import QuotationError, transition as move
    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("quotes.approve"):
        messages.error(request, "You do not have permission to approve quotations.")
        return redirect("web:quotation_detail", pk=pk)
    to_status = request.POST.get("status")
    try:
        move(quote, request.user, to_status=to_status,
             note=request.POST.get("note", ""))
    except QuotationError as exc:
        messages.error(request, str(exc))
    else:
        from apps.core.audit import audit
        audit(request, f"quotation.{to_status}", entity=quote)
        messages.success(request, f"{quote.display_number} → "
                                  f"{quote.get_status_display()}.")
    return redirect("web:quotation_detail", pk=pk)


@login_required
@require_POST
def quotation_revise(request, pk):
    from apps.quotes.services import create_revision
    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("quotes.create"):
        messages.error(request, "You do not have permission.")
        return redirect("web:quotation_detail", pk=pk)
    revised = create_revision(quote, request.user,
                              reason=request.POST.get("reason", ""))
    messages.success(request, f"Revision {revised.revision} created. The issued "
                              "version is unchanged.")
    return redirect("web:quotation_detail", pk=revised.id)


@login_required
@require_POST
def quotation_po(request, pk):
    """Capture the customer's purchase order from the uploaded document alone —
    the PO number, date, value and payment terms are read off the file by the
    shared Document Intelligence service, so nothing is typed."""
    from apps.knowledge.document_intelligence import (
        extract_po_fields,
        extract_text_from_upload,
    )
    from apps.quotes.services import QuotationError, record_purchase_order
    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("quotes.create"):
        messages.error(request, "You do not have permission.")
        return redirect("web:quotation_detail", pk=pk)
    f = request.FILES.get("document")
    if not f:
        messages.error(request, "Attach the purchase order document to save it.")
        return redirect("web:quotation_detail", pk=pk)
    from django.core.exceptions import ValidationError

    from apps.core.uploads import validate_upload
    try:
        validate_upload(f)
    except ValidationError as exc:
        messages.error(request, exc.messages[0])
        return redirect("web:quotation_detail", pk=pk)

    fields = extract_po_fields(extract_text_from_upload(f), company=quote.company,
                               user=request.user, use_ai=True)
    f.seek(0)                       # reading consumed the file; rewind before saving
    extracted_number = (fields.get("po_number") or "").strip()
    # If nothing at all could be read — no number, value, date or terms — the file
    # is almost certainly not a purchase order. Reject it rather than record a
    # bogus PO against the quotation.
    if not any([extracted_number, fields.get("value"), fields.get("po_date"),
                (fields.get("payment_terms") or "").strip()]):
        messages.error(
            request, f"“{f.name}” doesn’t look like a purchase order — nothing "
                     "could be read from it. Upload the customer’s PO document.")
        return redirect("web:quotation_detail", pk=pk)
    try:
        po = record_purchase_order(
            quote, request.user,
            po_number=extracted_number or f"PO-{quote.number}",
            value=_decimal_or_none(fields.get("value")),
            po_date=fields.get("po_date") or None,
            document=f,
            notes=(fields.get("payment_terms") or "").strip(),
        )
    except QuotationError as exc:
        messages.error(request, str(exc))
    else:
        from apps.core.audit import audit
        audit(request, "purchase_order.uploaded", entity=po)
        if extracted_number:
            messages.success(request, f"PO {po.po_number} saved — details read "
                                      f"from {f.name}.")
        else:
            # Some PO fields read, but not the number — save it, but flag that the
            # reference needs a human check rather than pretending it was read.
            messages.warning(
                request, f"Saved {f.name}, but the PO number couldn’t be read — "
                         f"recorded as {po.po_number}. Check it is the right "
                         "document and reference.")
        # Warn (do not block) if the same PO number is on another of this
        # customer's quotations — it may be a mis-keyed reference.
        from apps.quotes.models import CustomerPurchaseOrder
        if quote.customer_id and CustomerPurchaseOrder.objects.filter(
                quotation__customer_id=quote.customer_id,
                po_number__iexact=po.po_number).exclude(quotation_id=quote.id).exists():
            messages.warning(
                request, f"Heads up: PO {po.po_number} is also attached to another "
                         f"quotation for {quote.customer.display_name}.")
        # The PO is linked — automatically turn the quotation into operational
        # work (project · phases · tasks) and open the Work Details page. This is
        # the commercial → operational hand-off, with nothing re-entered.
        from apps.quotes.services import initiate_work_from_quotation
        try:
            project = initiate_work_from_quotation(quote, request.user)
        except QuotationError:
            project = None
        if project is not None:
            audit(request, "work.initiated", entity=project)
            messages.success(request, f"Work {project.number} created from "
                                      f"{quote.number} — opening it now.")
            return redirect("web:project_detail", pk=project.id)
    return redirect("web:quotation_detail", pk=pk)


@login_required
@require_POST
def quotation_start_work(request, pk):
    """Option 1 — start operational work directly from an approved quotation
    (no purchase order required). Creates the project · phases · tasks from the
    quotation and opens the Work Details page. Idempotent."""
    from apps.quotes.services import QuotationError, initiate_work_from_quotation
    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("projects.create"):
        messages.error(request, "You do not have permission to start work.")
        return redirect("web:quotation_detail", pk=pk)
    try:
        project = initiate_work_from_quotation(quote, request.user)
    except QuotationError as exc:
        messages.error(request, str(exc))
        return redirect("web:quotation_detail", pk=pk)
    from apps.core.audit import audit
    audit(request, "work.initiated", entity=project)
    messages.success(request, f"Work {project.number} created from {quote.number}.")
    return redirect("web:project_detail", pk=project.id)


@login_required
@require_POST
def quotation_award(request, pk):
    """Hand the quotation to execution. Not reversible, so it is confirmed."""
    from apps.quotes.services import QuotationError, award_to_work
    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("projects.create"):
        messages.error(request, "You do not have permission to award work.")
        return redirect("web:quotation_detail", pk=pk)
    try:
        result = award_to_work(quote, request.user,
                               work_name=request.POST.get("work_name", "").strip())
    except QuotationError as exc:
        messages.error(request, str(exc))
        return redirect("web:quotation_detail", pk=pk)

    project = result["project"]
    messages.success(
        request,
        f"Awarded. Work '{result['task'].name}' created"
        + (f" under {project.number}." if project else ".")
        + " This quotation is now read-only — changes are revisions.")
    return redirect("web:work_detail", pk=result["task"].id)


@login_required
@require_POST
def quotation_create_invoice(request, pk):
    """Raise a tax invoice from the quotation once a matching PO is linked."""
    from apps.quotes.services import QuotationError, create_invoice_document
    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("quotes.create"):
        messages.error(request, "You do not have permission.")
        return redirect("web:quotation_detail", pk=pk)
    try:
        doc = create_invoice_document(quote, request.user)
    except QuotationError as exc:
        messages.error(request, str(exc))
        return redirect("web:quotation_detail", pk=pk)
    from apps.core.audit import audit
    audit(request, "invoice.created", entity=doc)
    messages.success(request, f"Tax invoice {doc.number} created from {quote.number}.")
    return redirect("web:commercial_document_detail", pk=doc.id)


@login_required
@require_POST
def quotation_create_delivery(request, pk):
    """Raise a delivery note from the quotation once a matching PO is linked."""
    from apps.quotes.services import QuotationError, create_delivery_document
    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("quotes.create"):
        messages.error(request, "You do not have permission.")
        return redirect("web:quotation_detail", pk=pk)
    try:
        doc = create_delivery_document(
            quote, request.user,
            delivery_date=request.POST.get("delivery_date", ""),
            delivery_address=request.POST.get("delivery_address", ""),
            driver=request.POST.get("driver", ""),
            receiver_name=request.POST.get("receiver_name", ""),
            delivery_notes=request.POST.get("delivery_notes", ""))
    except QuotationError as exc:
        messages.error(request, str(exc))
        return redirect("web:quotation_detail", pk=pk)
    from apps.core.audit import audit
    audit(request, "delivery_note.created", entity=doc)
    messages.success(request, f"Delivery note {doc.number} created from {quote.number}.")
    return redirect("web:commercial_document_detail", pk=doc.id)


@login_required
@xframe_options_sameorigin
def commercial_document_pdf(request, pk):
    """Serve a generated tax invoice or delivery note PDF."""
    from apps.quotes.models import CommercialDocument
    from apps.quotes.pdf import delivery_note_pdf_bytes, invoice_pdf_bytes

    doc = get_object_or_404(CommercialDocument.objects.all(), pk=pk)
    inline = bool(request.GET.get("inline"))
    if not inline and not request.user.has_perm_code("quotes.download"):
        return HttpResponseForbidden("You do not have permission to download this.")
    if doc.kind == CommercialDocument.Kind.INVOICE:
        pdf = invoice_pdf_bytes(doc)
    else:
        pdf = delivery_note_pdf_bytes(doc)
    resp = HttpResponse(pdf, content_type="application/pdf")
    disposition = "inline" if inline else "attachment"
    resp["Content-Disposition"] = f'{disposition}; filename="{doc.number}.pdf"'
    if not inline:
        from apps.core.audit import audit
        audit(request, f"{doc.kind}.pdf_downloaded", entity=doc)
    return resp


#: The commercial life of an invoice / delivery note, for its timeline.
_COMDOC_STEPS = ["Draft", "Approved"]
_COMDOC_STAGE = {"draft": 0, "approved": 1, "finalized": 1, "sent": 1}


@login_required
def commercial_document_detail(request, pk):
    """Review workspace for a tax invoice or delivery note — the same interface
    as the quotation: PDF preview, status banner, timeline, lifecycle actions,
    read-only once finalized."""
    from apps.quotes.models import CommercialDocument
    from apps.quotes.services import (
        can_generate_documents,
        commercial_document_next_statuses,
    )

    doc = get_object_or_404(
        CommercialDocument.objects.select_related("quotation", "purchase_order"), pk=pk)
    can_quote = request.user.has_perm_code("quotes.create")
    stage = _COMDOC_STAGE.get(doc.status, 0)
    timeline = [{"label": label, "done": i <= stage, "current": i == stage}
                for i, label in enumerate(_COMDOC_STEPS)]

    return render(request, "web/commercial_document_detail.html", {
        "doc": doc,
        "quote": doc.quotation,
        "is_invoice": doc.kind == CommercialDocument.Kind.INVOICE,
        "can_view_money": _can_view_money(request.user),
        # Approve is the single, final step; there is no finalize or send.
        "can_approve": request.user.has_perm_code("quotes.approve") and doc.status == "draft",
        "can_download": doc.is_finalized and request.user.has_perm_code("quotes.download"),
        # From here you can also reach the sibling document off the same
        # quotation — view it if it exists, otherwise raise it.
        "can_generate_docs": can_quote and can_generate_documents(doc.quotation),
        "existing_invoice": doc.quotation.commercial_documents.filter(
            kind="invoice").first(),
        "existing_delivery": doc.quotation.commercial_documents.filter(
            kind="delivery").first(),
        "next_statuses": commercial_document_next_statuses(doc),
        "timeline": timeline,
    })


@login_required
@require_POST
def commercial_document_transition(request, pk):
    from apps.quotes.models import CommercialDocument
    from apps.quotes.services import QuotationError, transition_commercial_document
    doc = get_object_or_404(CommercialDocument.objects.all(), pk=pk)
    if not request.user.has_perm_code("quotes.approve"):
        messages.error(request, "You do not have permission to approve documents.")
        return redirect("web:commercial_document_detail", pk=pk)
    try:
        transition_commercial_document(doc, request.user,
                                       request.POST.get("status"))
    except QuotationError as exc:
        messages.error(request, str(exc))
    else:
        messages.success(request, f"{doc.number} → {doc.get_status_display()}.")
    return redirect("web:commercial_document_detail", pk=pk)


@login_required
def commercial_document_excel(request, pk):
    """Export a generated document's items to .xlsx — an invoice with prices, a
    delivery note with ordered quantities and no prices."""
    import io

    import openpyxl
    from openpyxl.styles import Font

    from apps.quotes.models import CommercialDocument
    if not request.user.has_perm_code("quotes.download"):
        return HttpResponseForbidden("You do not have permission to export this.")
    doc = get_object_or_404(CommercialDocument.objects.select_related("quotation"), pk=pk)
    quote = doc.quotation
    is_invoice = doc.kind == CommercialDocument.Kind.INVOICE

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice" if is_invoice else "Delivery note"
    ws.append([f"{doc.get_kind_display()} {doc.number}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([quote.client_name, "", "", f"Quotation {quote.number}"])
    ws.append([])
    if is_invoice:
        ws.append(["#", "Description", "Qty", "Unit", "Unit price", "Line total"])
    else:
        ws.append(["#", "Description", "Ordered", "Delivered", "Outstanding", "Unit"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for ln in quote.lines.all():
        if is_invoice:
            ws.append([ln.position, ln.description, float(ln.qty), ln.unit,
                       float(ln.effective_unit_price), float(ln.line_total)])
        else:
            ws.append([ln.position, ln.description, float(ln.qty), "", "", ln.unit])
    if is_invoice:
        ws.append([])
        ws.append(["", "", "", "", "Subtotal", float(quote.net_total)])
        ws.append(["", "", "", "", f"VAT ({quote.vat_rate:g}%)", float(quote.vat_amount)])
        ws.append(["", "", "", "", "Total", float(quote.invoice_total)])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    for col, width in {"A": 6, "B": 48, "C": 12, "D": 12, "E": 14, "F": 14}.items():
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{doc.number}.xlsx"'
    from apps.core.audit import audit
    audit(request, f"{doc.kind}.excel_exported", entity=doc)
    return resp


@login_required
def quotation_suggest(request, pk):
    """LulaAI proposes lines this quotation may be missing. Read-only."""
    from apps.quotes.estimating_ai import apply_suggestions, suggest_lines
    from apps.quotes.services import QuotationError

    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("quotes.create"):
        messages.error(request, "You do not have permission.")
        return redirect(_edit_url(pk))

    if request.method == "POST":
        # Recomputed server-side; the browser is trusted only for which items
        # were ticked.
        suggestion = suggest_lines(quote, request.user, use_ai=False)
        indexes = [i for i in request.POST.getlist("candidate") if i.isdigit()]
        try:
            created = apply_suggestions(quote, request.user, suggestion, indexes)
        except QuotationError as exc:
            messages.error(request, str(exc))
        else:
            messages.success(request, f"Added {created} line(s)." if created
                             else "Nothing selected — nothing was added.")
        return redirect(_edit_url(pk))

    return render(request, "web/quotation_suggest.html", {
        "quote": quote,
        "suggestion": suggest_lines(quote, request.user),
        "can_view_money": _can_view_money(request.user),
    })


@login_required
@require_POST
def quotation_template(request, pk):
    """Adopt the sections a job of this type is normally priced in."""
    from apps.quotes.services import apply_type_template
    quote, allowed = _quote_guard(request, pk)
    if allowed:
        created = apply_type_template(quote, request.user)
        messages.success(request, f"Added {created} section(s)." if created
                         else "Nothing to add — those sections already exist.")
    return redirect(_edit_url(pk))


@login_required
@require_POST
def quotation_line_move(request, pk):
    from apps.quotes.models import QuotationLine
    from apps.quotes.services import move_line
    quote, allowed = _quote_guard(request, pk)
    if allowed:
        line = get_object_or_404(QuotationLine.objects.filter(quotation=quote),
                                 pk=request.POST.get("line"))
        move_line(line, direction=request.POST.get("direction", "up"))
    return redirect(_edit_url(pk))


@login_required
@require_POST
def quotation_document(request, pk):
    """Attach the RFQ, drawings, a BOQ — whatever the quotation was built from."""
    from apps.quotes.models import QuotationDocument
    quote = get_object_or_404(Quotation.objects.all(), pk=pk)
    if not request.user.has_perm_code("quotes.create"):
        messages.error(request, "You do not have permission.")
        return redirect(_edit_url(pk))

    if request.POST.get("action") == "delete":
        doc = get_object_or_404(QuotationDocument.objects.filter(quotation=quote),
                                pk=request.POST.get("document"))
        doc.delete()
        messages.success(request, "Document removed.")
    elif request.FILES.get("file"):
        QuotationDocument.objects.create(
            company=quote.company, quotation=quote, file=request.FILES["file"],
            name=request.POST.get("name", "").strip() or request.FILES["file"].name,
            doc_type=request.POST.get("doc_type", "").strip(),
            created_by=request.user, updated_by=request.user)
        messages.success(request, "Document attached.")
    else:
        messages.error(request, "Choose a file to attach.")
    return redirect(_edit_url(pk))


@login_required
@require_POST
def quotation_lines_bulk(request, pk):
    """Add many items at once — typed into a grid, or pasted from a spreadsheet."""
    from apps.quotes.models import QuotationSection
    from apps.quotes.services import QuotationError, add_lines_bulk, parse_pasted_lines

    quote, allowed = _quote_guard(request, pk)
    if not allowed:
        return redirect(_edit_url(pk))

    section = QuotationSection.objects.filter(
        quotation=quote, pk=request.POST.get("section")).first()

    pasted = request.POST.get("pasted", "").strip()
    if pasted:
        rows = parse_pasted_lines(pasted)
    else:
        # Parallel arrays from the grid; blank descriptions are dropped, so an
        # estimator can leave spare rows empty without thinking about it.
        rows = []
        descriptions = request.POST.getlist("g_description")
        qtys = request.POST.getlist("g_qty")
        units = request.POST.getlist("g_unit")
        costs = request.POST.getlist("g_unit_cost")
        markups = request.POST.getlist("g_markup")
        for index, description in enumerate(descriptions):
            if not description.strip():
                continue
            rows.append({
                "description": description,
                "qty": _decimal_or_none(qtys[index] if index < len(qtys) else "")
                       or Decimal("1"),
                "unit": (units[index] if index < len(units) else "each") or "each",
                "unit_cost": _decimal_or_none(
                    costs[index] if index < len(costs) else "") or Decimal("0"),
                "markup_pct": _decimal_or_none(
                    markups[index] if index < len(markups) else "") or Decimal("0"),
            })

    if not rows:
        messages.error(request, "Nothing to add — every row was blank.")
        return redirect(_edit_url(pk))

    try:
        created = add_lines_bulk(quote, request.user, rows, section=section)
    except QuotationError as exc:
        messages.error(request, str(exc))
    else:
        messages.success(request, f"Added {created} item(s).")
    return redirect(_edit_url(pk))
